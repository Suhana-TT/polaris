#!/usr/bin/env python3
# SPDX-FileCopyrightText: (C) 2025 Tenstorrent AI ULC
# SPDX-License-Identifier: Apache-2.0

"""
Map TTNN hardware profiler table rows (CSV or Excel) to Polaris *opstats* ``opname`` values.

Polaris simulator stats (``HLMStats``) use graph node names such as ``ttsim.ttnn.Op_17``, not
``MatMul_0``-style names from :mod:`tools.profiling.profiler_to_polaris_converter`.

Typical workflow:

1. Export profiler ops to ``.csv`` or ``.xlsx`` (same column names as TTNN ops perf CSV).
2. Run Polaris on the same workload with stats CSV enabled.
3. Call :func:`map_profiler_rows_to_polaris_opnames` with both row lists (or use
   :func:`load_profiler_ops_table` / :func:`load_polaris_opstats_csv`).

Use :func:`profiler_polaris_layer_diff` or CLI ``--diff-only`` to list ops present only on the
hardware-profiler side or only in Polaris stats (with types and fingerprints).

Matching strategies:

* ``fingerprint`` (default): (``optype``, input shapes, output shapes) using the same optype and
  WZYX parsing rules as ``tools.profiling.profiler_to_polaris_converter``. Stable for duplicate op signatures by
  consuming Polaris ops in ``opnum`` order for each signature. By default
  ``use_layout_attr_shapes`` is True: Polaris layout ops use ``attrs`` (``TilizeWithValPadding``
  ``output_padded_shape``; ``UntilizeWithValUnpadding`` ``output_shape`` / ``output_tensor_end``), and
  the profiler side uses ``*_PAD[PADDED]`` columns when present (else ``LOGICAL``) for
  ``TilizeWithValPadding`` **output** and ``UntilizeWithValUnpadding`` **input_0**. Pass False or CLI
  ``--no-layout-attr-shapes`` for ``LOGICAL``-only profiler columns and Polaris tensor strings only.
* ``ordered``: align profiler rows sorted by ``GLOBAL CALL COUNT`` with Polaris rows sorted by
  ``opnum``; requires equal lengths.

Optional ``layer_key_column`` names an extra column in the profiler sheet (e.g. ``LAYER NAME``)
used as the human-readable *profiler* key in results; otherwise keys are synthetic.
"""

from __future__ import annotations

import argparse
import ast
import csv
import math
import re
import sys
import yaml
from collections import defaultdict, deque
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal, Mapping, Sequence

_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from tools.profiling.profiler_to_polaris_converter import (  # noqa: E402
    map_optype_to_polaris,
    parse_tensor_dimensions,
)


def _parse_profiler_attributes(raw: str | None) -> dict[str, Any]:
    if not raw or not str(raw).strip():
        return {}
    try:
        parsed = yaml.safe_load(str(raw).replace(';', ','))
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _parse_polaris_row_attrs(cell: str | None) -> dict[str, Any]:
    """Parse Polaris ``opstats`` ``attrs`` cell (often Python ``repr`` of a dict)."""
    s = (cell or '').strip()
    if not s or s == '{}':
        return {}
    try:
        v = ast.literal_eval(s)
        return v if isinstance(v, dict) else {}
    except (ValueError, SyntaxError, TypeError):
        pass
    try:
        parsed = yaml.safe_load(s.replace("'", '"'))
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _coerce_output_padded_shape(raw: Any) -> tuple[int, ...] | None:
    """
    Normalize ``attrs['output_padded_shape']`` from Polaris CSV.

    Simulator exports may use a Python list ``[8, 14, 32, 1024]`` or a string such as
    ``Shape([8, 224, 32, 64])``.
    """
    if raw is None:
        return None
    if isinstance(raw, (list, tuple)):
        if not raw:
            return None
        try:
            return _normalize_shape_tuple(tuple(int(x) for x in raw))
        except (TypeError, ValueError):
            return None
    if isinstance(raw, str):
        s = raw.strip()
        m = re.search(r'\[([^\]]+)\]', s)
        if not m:
            return None
        inner = m.group(1).replace('×', 'x')
        if 'x' in inner and ',' not in inner:
            parts = [p.strip() for p in inner.split('x') if p.strip()]
        else:
            parts = [p.strip() for p in inner.split(',') if p.strip()]
        if not parts:
            return None
        try:
            return _normalize_shape_tuple(tuple(int(p) for p in parts))
        except ValueError:
            return None
    return None


def _untilize_unpadding_logical_output_from_attrs(pattrs: Mapping[str, Any]) -> tuple[int, ...] | None:
    """
    True logical output shape for UntilizeWithValUnpadding from Polaris ``attrs``.

    Simulator ops store ``output_shape`` (dims). CSV exports may use ``output_tensor_end`` with
    last valid indices per dimension; output size is ``end + 1`` (see ``untilize_with_unpadding``).
    """
    raw_os = pattrs.get('output_shape')
    if raw_os is not None:
        t = _coerce_output_padded_shape(raw_os)
        if t is not None:
            return t
    raw_end = pattrs.get('output_tensor_end')
    if raw_end is None:
        return None
    ends = _coerce_output_padded_shape(raw_end)
    if ends is None:
        return None
    try:
        return _normalize_shape_tuple(tuple(int(e) + 1 for e in ends))
    except (TypeError, ValueError):
        return None


def _excel_cell_str(v: Any) -> str:
    if v is None:
        return ''
    if isinstance(v, float) and math.isnan(v):
        return ''
    s = str(v).strip()
    if s.lower() in ('nan', 'none', 'nat'):
        return ''
    return s


def _normalize_profiler_row_keys(row: Mapping[str, Any]) -> dict[str, str]:
    """Coerce header keys/values to str strips (Excel may use non-str types)."""
    out: dict[str, str] = {}
    for k, v in row.items():
        key = str(k).strip() if k is not None else ''
        out[key] = _excel_cell_str(v)
    return out


def _map_profiler_opcode_to_polaris_optype(opcode: str, attrs: dict[str, Any]) -> str:
    """Extend :func:`map_optype_to_polaris` for profiler / Excel opcodes (e.g. ``MatmulDeviceOperation``)."""
    oc = (opcode or '').strip()
    if oc == 'ReshapeViewDeviceOperation':
        return 'Reshape'
    if oc.endswith('DeviceOperation'):
        stem = oc[: -len('DeviceOperation')]
        # map_optype_to_polaris keys vary: some use *DeviceOperation*, some bare names.
        stem_to_lookup: dict[str, str] = {
            'Permute': 'PermuteDeviceOperation',
            'Softmax': 'SoftmaxDeviceOperation',
            'LayerNorm': 'LayerNormalization',
        }
        lookup = stem_to_lookup.get(stem, stem)
        return map_optype_to_polaris(lookup, attrs)
    return map_optype_to_polaris(oc, attrs)


def profiler_op_index(row: Mapping[str, str]) -> int:
    """0-based op index from ``GLOBAL CALL COUNT`` (1024, 2048, …) or row order."""
    gcc = row.get('GLOBAL CALL COUNT', '').strip()
    if gcc:
        try:
            return int(float(gcc)) // 1024 - 1
        except ValueError:
            pass
    return -1


def _squeeze_leading_ones(shape: tuple[int, ...]) -> tuple[int, ...]:
    """Match profiler WZYX (1,1,y,x) to Polaris tensor strings that omit leading 1 dims."""
    t = list(shape)
    while len(t) > 1 and t[0] == 1:
        t.pop(0)
    return tuple(t)


def _normalize_shape_tuple(shape: tuple[int, ...]) -> tuple[int, ...]:
    return _squeeze_leading_ones(shape)


def _shapes_from_polaris_tensor_field(field: str) -> tuple[tuple[int, ...], ...]:
    """
    Parse Polaris ``input_tensors`` / ``output_tensors`` cell:
    ``name[8x224x768]:float16;name2[768x768]:float16`` → ``((8,224,768), (768,768))``.
    """
    chunks = [c.strip() for c in field.split(';') if c.strip()]
    shapes: list[tuple[int, ...]] = []
    for chunk in chunks:
        m = re.search(r'\[([^\]]+)\]', chunk)
        if not m:
            continue
        inner = m.group(1).replace('×', 'x')
        shapes.append(tuple(int(x.strip()) for x in inner.split('x') if x.strip()))
    return tuple(_normalize_shape_tuple(s) for s in shapes)


def _optype_match_key(optype: str) -> str:
    """Polaris uses CamelCase (Add); profiler attrs often yield UPPER (ADD)."""
    return str(optype).strip().lower()


def polaris_op_signature(
    row: Mapping[str, str],
    *,
    use_layout_attr_shapes: bool = True,
) -> tuple[str, tuple[tuple[int, ...], ...], tuple[tuple[int, ...], ...]]:
    """
    Build fingerprint tuple for a Polaris opstats row.

    Parameters
    ----------
    use_layout_attr_shapes
        Default True. When True, refine fingerprints from ``attrs`` for:

        * ``TilizeWithValPadding``: **output** from ``output_padded_shape`` (tile extent).
        * ``UntilizeWithValUnpadding``: **output** from ``output_shape`` or ``output_tensor_end``.

        Set False to use only ``input_tensors`` / ``output_tensors`` strings for those ops.
    """
    optype = _optype_match_key(str(row.get('optype', '')).strip())
    in_shapes = _shapes_from_polaris_tensor_field(row.get('input_tensors', '') or '')
    out_shapes = _shapes_from_polaris_tensor_field(row.get('output_tensors', '') or '')
    if use_layout_attr_shapes:
        pattrs = _parse_polaris_row_attrs(row.get('attrs'))
        if optype == 'tilizewithvalpadding':
            tup = _coerce_output_padded_shape(pattrs.get('output_padded_shape'))
            if tup is not None:
                out_shapes = (tup,)
        elif optype in ('untilizewithvalunpadding', 'untilizewithunpadding'):
            uo = _untilize_unpadding_logical_output_from_attrs(pattrs)
            if uo is not None:
                out_shapes = (uo,)
    return (optype, in_shapes, out_shapes)


def profiler_op_signature(
    row: Mapping[str, str],
    *,
    use_layout_attr_shapes: bool = True,
) -> tuple[str, tuple[tuple[int, ...], ...], tuple[tuple[int, ...], ...]]:
    """
    Hardware-profiler fingerprint; same tuple shape as :func:`polaris_op_signature`.

    When ``use_layout_attr_shapes`` is True, uses ``*_PAD[PADDED]`` via
    :func:`tools.profiling.profiler_to_polaris_converter.parse_tensor_dimensions` for ``TilizeWithValPadding``
    **output_0** and ``UntilizeWithValUnpadding`` **input_0** when those columns exist; otherwise
    ``*_PAD[LOGICAL]``. When False, always ``LOGICAL``.
    """
    attrs = _parse_profiler_attributes(row.get('ATTRIBUTES'))
    opcode = row.get('OP CODE', '') or ''
    optype = _optype_match_key(_map_profiler_opcode_to_polaris_optype(opcode, attrs))
    row_d: dict[str, Any] = dict(row)
    ins: list[tuple[int, ...]] = []
    for idx in (0, 1, 2):
        t = None
        if use_layout_attr_shapes and optype in ('untilizewithvalunpadding', 'untilizewithunpadding') and idx == 0:
            t = parse_tensor_dimensions(row_d, 'INPUT', idx, pad_extent='PADDED')
        if t is None:
            t = parse_tensor_dimensions(row_d, 'INPUT', idx)
        if t:
            ins.append(_normalize_shape_tuple(tuple(t['dims'])))
    outs: list[tuple[int, ...]] = []
    t0 = None
    if use_layout_attr_shapes and optype == 'tilizewithvalpadding':
        t0 = parse_tensor_dimensions(row_d, 'OUTPUT', 0, pad_extent='PADDED')
    if t0 is None:
        t0 = parse_tensor_dimensions(row_d, 'OUTPUT', 0)
    if t0:
        outs.append(_normalize_shape_tuple(tuple(t0['dims'])))
    return (optype, tuple(ins), tuple(outs))


def profiler_layer_key(
    row: Mapping[str, str],
    row_index: int,
    *,
    layer_key_column: str | None,
) -> str:
    if layer_key_column:
        k = layer_key_column.strip()
        if k in row and str(row[k]).strip():
            return str(row[k]).strip()
    gcc = row.get('GLOBAL CALL COUNT', '').strip()
    op = row.get('OP CODE', '').strip()
    if gcc:
        return f'{op}@{gcc}'
    return f'{op}#row{row_index}'


@dataclass(frozen=True)
class ProfilerPolarisOpnameMapping:
    """One profiler row aligned with one Polaris stats row."""

    profiler_layer_key: str
    polaris_opname: str
    profiler_op_index: int
    polaris_opnum: int
    optype: str
    strategy: str


@dataclass(frozen=True)
class ProfilerOnlyLayer:
    """Hardware-profiler row with no matching Polaris op (same fingerprint strategy)."""

    layer_key: str
    profiler_op_code: str
    mapped_optype: str
    signature: str
    profiler_row_index: int


@dataclass(frozen=True)
class PolarisOnlyLayer:
    """Polaris stats op with no matching hardware-profiler row."""

    polaris_opname: str
    polaris_opnum: int
    optype: str
    signature: str


@dataclass(frozen=True)
class ProfilerPolarisLayerDiff:
    """Layers present on one side only, after fingerprint pairing."""

    only_in_profiler: tuple[ProfilerOnlyLayer, ...]
    only_in_polaris: tuple[PolarisOnlyLayer, ...]
    matched_count: int


def _format_signature(sig: tuple[Any, ...]) -> str:
    return repr(sig)


def profiler_polaris_layer_diff(
    profiler_rows: Sequence[Mapping[str, Any]],
    polaris_opstats_rows: Sequence[Mapping[str, Any]],
    *,
    layer_key_column: str | None = None,
    skip_polaris_fused: bool = False,
    use_layout_attr_shapes: bool = True,
) -> ProfilerPolarisLayerDiff:
    """
    After fingerprint matching, list ops that appear only on the profiler side or only in Polaris.

    Uses the same pairing rules as :func:`map_profiler_rows_to_polaris_opnames` with
    ``strategy='fingerprint'`` and ``allow_unmatched=True``. Unmatched profiler rows and any
    Polaris queue entries left over are classified as unpaired.

    Parameters
    ----------
    use_layout_attr_shapes
        Same as :func:`map_profiler_rows_to_polaris_opnames`.

    Raises
    ------
    ValueError
        If ``profiler_rows`` contains a row with no ``OP CODE`` / empty signature (wrong Excel sheet).
    """
    prof_sorted = _sorted_profiler_rows(profiler_rows)
    pol = _sorted_polaris_rows(polaris_opstats_rows)
    if skip_polaris_fused:
        pol = [r for r in pol if str(r.get('fused', '')).lower() not in ('true', '1', 'yes')]

    queues: dict[tuple[Any, ...], deque[tuple[str, str, int]]] = defaultdict(deque)
    for pl in pol:
        sig = polaris_op_signature(pl, use_layout_attr_shapes=use_layout_attr_shapes)
        try:
            opn = int(str(pl.get('opnum', 0)))
        except ValueError:
            opn = 0
        queues[sig].append(
            (str(pl.get('opname', '')), str(pl.get('optype', '')), opn),
        )

    only_prof: list[ProfilerOnlyLayer] = []
    matched = 0
    for j, (_, pr) in enumerate(prof_sorted):
        sig = profiler_op_signature(pr, use_layout_attr_shapes=use_layout_attr_shapes)
        pidx = profiler_op_index(pr)
        if pidx < 0:
            pidx = j
        op_code = pr.get('OP CODE', '').strip()
        mapped_ot = _map_profiler_opcode_to_polaris_optype(
            op_code, _parse_profiler_attributes(pr.get('ATTRIBUTES'))
        )
        key = profiler_layer_key(pr, j, layer_key_column=layer_key_column)
        if j == 0 and not op_code and not sig[0]:
            raise ValueError(
                'First profiler row has empty OP CODE and signature — wrong Excel sheet? '
                'Pass sheet_name=... to load_profiler_ops_table (e.g. "vitfull").'
            )
        q = queues[sig]
        if not q:
            only_prof.append(
                ProfilerOnlyLayer(
                    layer_key=key,
                    profiler_op_code=op_code,
                    mapped_optype=mapped_ot,
                    signature=_format_signature(sig),
                    profiler_row_index=j,
                )
            )
            continue
        q.popleft()
        matched += 1

    only_pol: list[PolarisOnlyLayer] = []
    for sig, dq in queues.items():
        while dq:
            opname, display_ot, opn = dq.popleft()
            only_pol.append(
                PolarisOnlyLayer(
                    polaris_opname=opname,
                    polaris_opnum=opn,
                    optype=display_ot,
                    signature=_format_signature(sig),
                )
            )
    only_pol.sort(key=lambda x: (x.polaris_opnum, x.polaris_opname))

    return ProfilerPolarisLayerDiff(
        only_in_profiler=tuple(only_prof),
        only_in_polaris=tuple(only_pol),
        matched_count=matched,
    )


def summarize_layer_diff_by_optype(diff: ProfilerPolarisLayerDiff) -> tuple[dict[str, int], dict[str, int]]:
    """Return (profiler_only_counts_by_mapped_optype, polaris_only_counts_by_optype)."""
    pc: dict[str, int] = {}
    for x in diff.only_in_profiler:
        pc[x.mapped_optype] = pc.get(x.mapped_optype, 0) + 1
    pl: dict[str, int] = {}
    for x in diff.only_in_polaris:
        pl[x.optype] = pl.get(x.optype, 0) + 1
    return pc, pl


def _sorted_polaris_rows(rows: Sequence[Mapping[str, str]]) -> list[dict[str, str]]:
    def opnum_key(r: Mapping[str, str]) -> int:
        try:
            return int(str(r.get('opnum', 0)))
        except ValueError:
            return 0

    return sorted((_normalize_profiler_row_keys(r) for r in rows), key=opnum_key)


def _sorted_profiler_rows(rows: Sequence[Mapping[str, str]]) -> list[tuple[int, dict[str, str]]]:
    indexed: list[tuple[int, dict[str, str]]] = []
    for i, r in enumerate(rows):
        d = _normalize_profiler_row_keys(r)
        pidx = profiler_op_index(d)
        indexed.append((pidx if pidx >= 0 else i, d))
    indexed.sort(key=lambda x: (x[0], x[1].get('GLOBAL CALL COUNT', '')))
    return indexed


def map_profiler_rows_to_polaris_opnames(
    profiler_rows: Sequence[Mapping[str, Any]],
    polaris_opstats_rows: Sequence[Mapping[str, Any]],
    *,
    strategy: Literal['fingerprint', 'ordered'] = 'fingerprint',
    layer_key_column: str | None = None,
    skip_polaris_fused: bool = False,
    allow_unmatched: bool = False,
    use_layout_attr_shapes: bool = True,
) -> list[ProfilerPolarisOpnameMapping]:
    """
    Pair each hardware-profiler op row with the corresponding Polaris ``opname``.

    Parameters
    ----------
    profiler_rows
        Rows from TTNN ops perf CSV / Excel (see repo ``csvs/ops_perf_results_*.csv``).
    polaris_opstats_rows
        Rows from Polaris ``*opstats.csv`` (must include ``opnum``, ``opname``, ``optype``,
        ``input_tensors``, ``output_tensors``).
    strategy
        ``fingerprint`` (default) or ``ordered`` (strict position after sort).
    layer_key_column
        Optional profiler column for human-readable keys (e.g. ``LAYER NAME``).
    skip_polaris_fused
        If True, drop Polaris rows where ``fused`` is truthy (``True`` / ``1``) before matching.
    allow_unmatched
        If True, profiler rows with no matching Polaris op (e.g. different workload scope) get
        empty ``polaris_opname`` and ``polaris_opnum=-1`` instead of raising.
    use_layout_attr_shapes
        Default True. See :func:`polaris_op_signature` and :func:`profiler_op_signature`.
    """
    prof_sorted = _sorted_profiler_rows(profiler_rows)
    pol = _sorted_polaris_rows(polaris_opstats_rows)
    if skip_polaris_fused:
        pol = [r for r in pol if str(r.get('fused', '')).lower() not in ('true', '1', 'yes')]

    if strategy == 'ordered':
        if len(prof_sorted) != len(pol):
            raise ValueError(
                f'ordered strategy requires equal row counts, got profiler={len(prof_sorted)} '
                f'polaris={len(pol)}'
            )
        out: list[ProfilerPolarisOpnameMapping] = []
        for j, ((_, pr), pl) in enumerate(zip(prof_sorted, pol)):
            pidx = profiler_op_index(pr)
            if pidx < 0:
                pidx = j
            out.append(
                ProfilerPolarisOpnameMapping(
                    profiler_layer_key=profiler_layer_key(pr, j, layer_key_column=layer_key_column),
                    polaris_opname=str(pl.get('opname', '')),
                    profiler_op_index=pidx,
                    polaris_opnum=int(pl.get('opnum', 0) or 0),
                    optype=str(pl.get('optype', '')),
                    strategy='ordered',
                )
            )
        return out

    # fingerprint: queue polaris (opname, display optype, opnum) per signature
    queues: dict[tuple[Any, ...], deque[tuple[str, str, int]]] = defaultdict(deque)
    for pl in pol:
        sig = polaris_op_signature(pl, use_layout_attr_shapes=use_layout_attr_shapes)
        try:
            opn = int(str(pl.get('opnum', 0)))
        except ValueError:
            opn = 0
        queues[sig].append(
            (str(pl.get('opname', '')), str(pl.get('optype', '')), opn),
        )

    out = []
    for j, (_, pr) in enumerate(prof_sorted):
        sig = profiler_op_signature(pr, use_layout_attr_shapes=use_layout_attr_shapes)
        pidx = profiler_op_index(pr)
        if pidx < 0:
            pidx = j
        q = queues[sig]
        if not q:
            if allow_unmatched:
                out.append(
                    ProfilerPolarisOpnameMapping(
                        profiler_layer_key=profiler_layer_key(pr, j, layer_key_column=layer_key_column),
                        polaris_opname='',
                        profiler_op_index=pidx,
                        polaris_opnum=-1,
                        optype=_map_profiler_opcode_to_polaris_optype(
                            pr.get('OP CODE', ''), _parse_profiler_attributes(pr.get('ATTRIBUTES'))
                        ),
                        strategy='unmatched',
                    )
                )
                continue
            raise ValueError(
                f'No Polaris op left for profiler row {j} key={profiler_layer_key(pr, j, layer_key_column=layer_key_column)!r} '
                f'signature={sig!r}'
            )
        opname, display_optype, opn = q.popleft()
        out.append(
            ProfilerPolarisOpnameMapping(
                profiler_layer_key=profiler_layer_key(pr, j, layer_key_column=layer_key_column),
                polaris_opname=opname,
                profiler_op_index=pidx,
                polaris_opnum=opn,
                optype=display_optype,
                strategy='fingerprint',
            )
        )
    return out


def load_polaris_opstats_csv(path: str | Path) -> list[dict[str, str]]:
    p = Path(path)
    with p.open(newline='', encoding='utf-8') as f:
        return [_normalize_profiler_row_keys(r) for r in csv.DictReader(f)]


def load_profiler_ops_table(path: str | Path, *, sheet_name: str | int | None = 0) -> list[dict[str, str]]:
    """
    Load profiler ops from ``.csv`` or ``.xlsx``.

    Excel support requires ``openpyxl`` (``pip install openpyxl``) or ``pandas``.
    """
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(p)

    if p.suffix.lower() == '.csv':
        with p.open(newline='', encoding='utf-8') as f:
            return [_normalize_profiler_row_keys(r) for r in csv.DictReader(f)]

    if p.suffix.lower() in ('.xlsx', '.xlsm'):
        try:
            import pandas as pd  # type: ignore[import-not-found]

            df = pd.read_excel(p, sheet_name=sheet_name)
            return [_normalize_profiler_row_keys(r) for r in df.to_dict('records')]
        except ImportError:
            pass
        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except ImportError as e:
            raise ImportError(
                'Reading .xlsx requires openpyxl or pandas. Install with: pip install openpyxl'
            ) from e
        wb = load_workbook(p, read_only=True, data_only=True)
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        elif isinstance(sheet_name, str):
            ws = wb[sheet_name]
        else:
            ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return []
        headers = [str(h).strip() if h is not None else '' for h in header]
        out: list[dict[str, str]] = []
        for tup in rows:
            if tup is None or all(v is None for v in tup):
                continue
            row = {headers[i]: ('' if tup[i] is None else str(tup[i]).strip()) for i in range(len(headers)) if headers[i]}
            out.append(row)
        return out

    raise ValueError(f'Unsupported profiler file type: {p.suffix}')


def mapping_to_dict(m: ProfilerPolarisOpnameMapping) -> dict[str, Any]:
    return {
        'profiler_layer_key': m.profiler_layer_key,
        'polaris_opname': m.polaris_opname,
        'profiler_op_index': m.profiler_op_index,
        'polaris_opnum': m.polaris_opnum,
        'optype': m.optype,
        'strategy': m.strategy,
    }


def profiler_keys_to_polaris_opnames(
    profiler_rows: Sequence[Mapping[str, Any]],
    polaris_opstats_rows: Sequence[Mapping[str, Any]],
    **kwargs: Any,
) -> dict[str, str]:
    """
    Convenience wrapper: ``{profiler_layer_key: polaris_opname}`` for matched rows only.

    Pass the same ``kwargs`` as :func:`map_profiler_rows_to_polaris_opnames`.
    """
    d: dict[str, str] = {}
    for m in map_profiler_rows_to_polaris_opnames(profiler_rows, polaris_opstats_rows, **kwargs):
        if m.polaris_opname:
            d[m.profiler_layer_key] = m.polaris_opname
    return d


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        'profiler_table',
        nargs='?',
        default='vitfull.xlsx',
        help='Profiler CSV or XLSX (default: vitfull.xlsx in cwd)',
    )
    ap.add_argument('polaris_opstats', help='Polaris *opstats.csv')
    ap.add_argument(
        '--strategy',
        choices=('fingerprint', 'ordered'),
        default='fingerprint',
        help='Matching strategy (default: fingerprint)',
    )
    ap.add_argument(
        '--layer-key-column',
        default=None,
        help='Profiler column to use as layer key (e.g. LAYER NAME)',
    )
    ap.add_argument(
        '--sheet',
        default=0,
        help='Excel sheet name or 0-based index (default: 0)',
    )
    ap.add_argument(
        '--skip-fused',
        action='store_true',
        help='Ignore Polaris rows with fused=True',
    )
    ap.add_argument(
        '--allow-unmatched',
        action='store_true',
        help='Leave polaris_opname empty when profiler op has no Polaris counterpart',
    )
    ap.add_argument(
        '--no-layout-attr-shapes',
        dest='use_layout_attr_shapes',
        action='store_false',
        help=(
            'Layout ops: Polaris tensor strings only; profiler *_PAD[LOGICAL] only (ignore attrs '
            'and *_PAD[PADDED])'
        ),
    )
    ap.set_defaults(use_layout_attr_shapes=True)
    ap.add_argument('--json', action='store_true', help='Print JSON lines')
    ap.add_argument(
        '--diff-only',
        action='store_true',
        help='Print layers only in profiler vs only in Polaris (with types), then exit',
    )
    args = ap.parse_args()

    prof_path = Path(args.profiler_table)
    pol_path = Path(args.polaris_opstats)

    prof = load_profiler_ops_table(prof_path, sheet_name=args.sheet)
    pol = load_polaris_opstats_csv(pol_path)

    if args.diff_only:
        diff = profiler_polaris_layer_diff(
            prof,
            pol,
            layer_key_column=args.layer_key_column,
            skip_polaris_fused=args.skip_fused,
            use_layout_attr_shapes=args.use_layout_attr_shapes,
        )
        pc, pl = summarize_layer_diff_by_optype(diff)
        print(f'Matched pairs (fingerprint): {diff.matched_count}')
        print()
        print(f'=== Only in profiler (hardware), {len(diff.only_in_profiler)} ops ===')
        for x in diff.only_in_profiler:
            print(
                f'{x.layer_key}\t{x.profiler_op_code}\t{x.mapped_optype}\t{x.signature}',
            )
        print()
        print('--- Count by mapped optype (profiler-only) ---')
        for k in sorted(pc.keys(), key=lambda s: (-pc[s], s)):
            print(f'  {k}: {pc[k]}')
        print()
        print(f'=== Only in Polaris (simulator), {len(diff.only_in_polaris)} ops ===')
        for x in diff.only_in_polaris:
            print(f'{x.polaris_opname}\topnum={x.polaris_opnum}\t{x.optype}\t{x.signature}')
        print()
        print('--- Count by optype (Polaris-only) ---')
        for k in sorted(pl.keys(), key=lambda s: (-pl[s], s)):
            print(f'  {k}: {pl[k]}')
        return 0

    maps = map_profiler_rows_to_polaris_opnames(
        prof,
        pol,
        strategy=args.strategy,
        layer_key_column=args.layer_key_column,
        skip_polaris_fused=args.skip_fused,
        allow_unmatched=args.allow_unmatched,
        use_layout_attr_shapes=args.use_layout_attr_shapes,
    )
    if args.json:
        import json

        for m in maps:
            print(json.dumps(mapping_to_dict(m)))
    else:
        for m in maps:
            print(f'{m.profiler_layer_key}\t{m.polaris_opname}\topnum={m.polaris_opnum}\t{m.optype}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
