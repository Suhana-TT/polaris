#!/usr/bin/env python
# SPDX-FileCopyrightText: (C) 2025 Tenstorrent AI ULC
# SPDX-License-Identifier: Apache-2.0

"""Compare Polaris and Profiler CSV layer sequences."""

import sys
import os
import argparse
import csv
from collections import defaultdict
from pathlib import Path
from typing import Any, DefaultDict, Dict, List, Optional, Tuple, Union
from dataclasses import dataclass

try:
    from op_canonical import normalize_polaris_optype, to_comparison_group  # type: ignore[import-not-found]
    from shape_canonical import (  # type: ignore[import-not-found]
        parse_shape_string as parse_shape,
        normalize_shape,
        compare_tensor_shapes,
        validate_binary_compatibility,
        validate_reshape_compatibility,
        compare_tensor_attributes,
    )
except ImportError:
    from .op_canonical import normalize_polaris_optype, to_comparison_group  # type: ignore
    from .shape_canonical import (  # type: ignore
        parse_shape_string as parse_shape,
        normalize_shape,
        compare_tensor_shapes,
        validate_binary_compatibility,
        validate_reshape_compatibility,
        compare_tensor_attributes,
    )

# Maximum distance to search forward for matching operations
DEFAULT_MAX_SEARCH_DISTANCE = 10

# Import layer extraction functions
try:
    from show_layers_polaris import layers_polaris  # type: ignore[import-not-found]
    from show_layers_profiler import layers_profiler  # type: ignore[import-not-found]
except ImportError:
    # Try relative imports
    from .show_layers_polaris import layers_polaris  # type: ignore
    from .show_layers_profiler import layers_profiler  # type: ignore


def sanitize_file_path(filepath: str) -> Path:
    """
    Sanitize and validate a user-provided file path.

    Resolves the path to an absolute path and validates that it:
    - Exists
    - Is a regular file (not a directory or special file)
    - Has an allowed file extension (.csv, .txt, .log, .tsv)
    - Is within a safe base directory (POLARIS_BASE_DIR env var, defaults to HOME)

    Security Note: This is a development/profiling tool that reads CSV files.
    Access is restricted to files within the base directory (user's home by default).
    Set POLARIS_BASE_DIR environment variable to use a different base directory.

    Args:
        filepath: User-provided file path string

    Returns:
        Resolved absolute Path object

    Raises:
        ValueError: If the path is invalid, doesn't exist, or violates security constraints
    """
    # Allowed file extensions for CSV/text files
    ALLOWED_EXTENSIONS = {'.csv', '.txt', '.log', '.tsv'}

    # === STEP 1: Pre-validation of user input ===
    # Check for malicious patterns before any path operations
    if not filepath or not isinstance(filepath, str):
        raise ValueError("File path must be a non-empty string")

    # Check for null bytes (path traversal attack vector)
    if '\0' in filepath:
        raise ValueError("File path contains null bytes")

    # Check for excessively long paths (potential DoS)
    # Using hardcoded 4096 (typical POSIX PATH_MAX) rather than os.pathconf
    # for simplicity and portability. This is a pre-validation security check;
    # the actual OS limit will be enforced by Path.resolve() anyway.
    if len(filepath) > 4096:
        raise ValueError("File path exceeds maximum allowed length")

    # Input validation complete - filepath is now safe to process
    validated_input = filepath

    # === STEP 2: Establish safe base directory ===
    # Determine and validate base directory before processing user filepath
    base_dir_env = os.environ.get('POLARIS_BASE_DIR') or os.getenv('HOME')
    if not base_dir_env:
        raise ValueError("Cannot determine safe base directory (HOME not set)")

    try:
        base_dir = Path(base_dir_env).resolve(strict=True)
        if not base_dir.is_dir():
            raise ValueError(f"Base directory is not a directory: {base_dir}")
    except (FileNotFoundError, RuntimeError, OSError) as e:
        raise ValueError(f"Invalid base directory: {base_dir_env}") from e

    # === STEP 3: Resolve and validate file path ===
    try:
        # Convert validated input to Path and resolve to absolute path
        # This resolves symlinks and .. references
        resolved_path = Path(validated_input).resolve(strict=True)

        # Verify it's a regular file (not a directory or special file)
        if not resolved_path.is_file():
            raise ValueError(f"Path is not a regular file: {validated_input}")

        # Check file extension against allowlist
        if resolved_path.suffix.lower() not in ALLOWED_EXTENSIONS:
            raise ValueError(
                f"File must have one of these extensions: {', '.join(ALLOWED_EXTENSIONS)}. "
                f"Got: {resolved_path.suffix or '(no extension)'}"
            )

        # === STEP 4: Enforce base directory restriction ===
        # Verify resolved path is within the safe base directory
        try:
            resolved_path.relative_to(base_dir)
        except ValueError:
            raise ValueError(
                f"Access denied: File must be within {base_dir}. "
                f"Attempted to access: {resolved_path}"
            )

        # All security checks passed - return sanitized path
        return resolved_path

    except FileNotFoundError:
        raise ValueError(f"File not found: {validated_input}")
    except RuntimeError as e:
        # Can occur with symlink loops
        raise ValueError(f"Invalid file path (possible symlink loop): {validated_input}") from e
    except OSError as e:
        raise ValueError(f"Cannot access file: {validated_input}") from e


@dataclass
class ComparisonStats:
    """Statistics for layer comparison."""
    total_matches: int = 0
    name_mismatches: int = 0
    shape_mismatches: int = 0
    input_shape_mismatches: int = 0
    output_shape_mismatches: int = 0
    attr_mismatches: int = 0
    unmatched_polaris: int = 0
    unmatched_profiler: int = 0
    ambiguous: int = 0


def parse_args() -> argparse.Namespace:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='Compare Polaris and Profiler CSV layer sequences'
    )
    parser.add_argument('file1', type=str, help='First CSV file (polaris or profiler)')
    parser.add_argument('file2', type=str, nargs='?', default=None,
                        help='Second CSV file (optional; required for shape comparison)')
    parser.add_argument(
        '--perf',
        action='store_true',
        help='Enable performance matching: show network-total and layer-type-wise '
             'duration comparison (ms). With two files, shows gap w.r.t. profiler. '
             'With one file, shows standalone breakdown.'
    )
    parser.add_argument(
        '--max-search-distance',
        type=int,
        default=DEFAULT_MAX_SEARCH_DISTANCE,
        help=f'Maximum distance to search forward for matching operations (default: {DEFAULT_MAX_SEARCH_DISTANCE})'
    )
    parser.add_argument(
        '--strip-leading-ones',
        action=argparse.BooleanOptionalAction,
        default=True,
        help='Strip all leading 1s from shapes (default: enabled). '
             'Leading 1s are a batch-dimension convention difference between Polaris and HW. '
             'Use --no-strip-leading-ones for strict matching.'
    )
    # NOTE: --strip-singleton-dims is currently required for fused head ops
    # (CreateQKVHeads, ConcatHeads) because HW uses 4D shapes with a
    # seq_groups=1 singleton dim (e.g. [B, 1, S, H]) while Polaris emits 3D
    # shapes (e.g. [B, S, H]). Additionally, HW implicitly reinterprets the
    # output of ConcatHeads from [B, 1, S, H] to [1, B, S, H] for
    # downstream ops without an explicit reshape. Future work: update the
    # Polaris shim to emit 4D shapes and model this implicit view change,
    # which would allow removing this flag for those ops.
    parser.add_argument(
        '--strip-singleton-dims',
        action='store_true',
        help='Strip all singleton (=1) dimensions from shapes regardless of position '
             '(handles HW seq_groups=1 convention)'
    )
    parser.add_argument(
        '--filter-optype',
        type=str,
        default=None,
        help='Filter to only compare layers with this operation type (case-insensitive)'
    )
    parser.add_argument(
        '--ignore-attrs',
        action='store_true',
        help='Skip tensor attribute comparison (dtype, layout, memory); compare shapes only'
    )
    parser.add_argument(
        '--summarize-by-signature',
        action='store_true',
        help='Print a rollup table keyed by layer optype (CSV name) plus normalized '
             'input/output shape signature. Uses the same --strip-leading-ones / '
             '--strip-singleton-dims rules as shape comparison. With two CSVs, prints '
             'one table per file. With --perf, adds summed duration (and Polaris LUT '
             'hits when available), and the profiler-vs-Polaris performance comparison '
             'is grouped by type+signature instead of by optype alone.',
    )
    parser.add_argument(
        '--xlsx',
        type=str,
        default=None,
        metavar='PATH',
        help='Also write an .xlsx report with three sheets: '
             '"Summary" (network-wide totals + shape/attr counts), '
             '"By Layer Type" (per canonical optype), and '
             '"By Layer Signature" (per optype + normalized in/out shape signature). '
             'In two-file mode each sheet includes Profiler vs Polaris columns and gap; '
             'in single-file mode only that source\'s counts/ms are emitted. '
             'Requires openpyxl (already a polarisdev dep).',
    )
    return parser.parse_args()


def normalize_optype(optype: str) -> str:
    """Normalize operation type name to canonical form then coarsen for
    sequence matching (e.g. add/mul/sub → binary).

    Delegates to :mod:`op_canonical` for the canonical form and comparison
    group mapping.
    """
    return to_comparison_group(normalize_polaris_optype(optype))


def detect_file_type(filepath: str) -> Optional[str]:
    """
    Detect if file is polaris or profiler CSV.

    Args:
        filepath: User-provided file path (will be sanitized)

    Returns:
        'polaris' if archname column found
        'profiler' if OP CODE column found
        None if neither found
    """
    try:
        # Sanitize the file path to prevent path traversal attacks
        safe_path = sanitize_file_path(filepath)

        with open(safe_path, 'r') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            if 'archname' in headers:
                return 'polaris'
            elif 'OP CODE' in headers:
                return 'profiler'
            return None
    except ValueError as e:
        # Path sanitization failed
        print(f"Invalid file path {filepath}: {e}", file=sys.stderr)
        return None
    except Exception as e:
        print(f"Error reading {filepath}: {e}", file=sys.stderr)
        return None




def find_next_match(
    layers: List[Dict[str, Any]],
    start_idx: int,
    target_optype: str,
    max_distance: Optional[int] = None
) -> Optional[int]:
    """
    Find the next occurrence of target_optype in layers starting from start_idx.

    Args:
        layers: List of layer dictionaries
        start_idx: Index to start searching from
        target_optype: Normalized operation type to search for
        max_distance: Maximum number of operations to search ahead (None = unlimited)

    Returns:
        Index of match, or None if not found within max_distance
    """
    end_idx = len(layers)
    if max_distance is not None:
        end_idx = min(end_idx, start_idx + max_distance)

    for i in range(start_idx, end_idx):
        if normalize_optype(layers[i]['optype']) == target_optype:
            return i
    return None


def format_shapes(shapes: List[str]) -> str:
    """Format list of shape strings for display."""
    if not shapes:
        return "[]"
    return "[" + ", ".join(shapes) + "]"


def compare_layers(
    polaris_layers: List[Dict[str, Any]],
    profiler_layers: List[Dict[str, Any]],
    max_search_distance: int = DEFAULT_MAX_SEARCH_DISTANCE,
    strip_leading_ones: bool = False,
    strip_singleton_dims: bool = False,
    ignore_attrs: bool = False,
) -> ComparisonStats:
    """
    Compare two layer sequences and print results.

    Returns:
        ComparisonStats object with statistics
    """
    stats = ComparisonStats()
    ndx_polaris = 0
    ndx_profiler = 0

    while ndx_polaris < len(polaris_layers) or ndx_profiler < len(profiler_layers):
        # Case 3: One sequence exhausted
        if ndx_polaris >= len(polaris_layers):
            # Profiler has remaining entries
            layer = profiler_layers[ndx_profiler]
            print(f"⊘ [F:{layer['seqno']}] {layer['optype']} (skipped in polaris)")
            stats.unmatched_profiler += 1
            ndx_profiler += 1
            continue

        if ndx_profiler >= len(profiler_layers):
            # Polaris has remaining entries
            layer = polaris_layers[ndx_polaris]
            print(f"⊘ [P:{layer['seqno']}] {layer['optype']} (skipped in profiler)")
            stats.unmatched_polaris += 1
            ndx_polaris += 1
            continue

        # Get current layers
        p_layer = polaris_layers[ndx_polaris]
        f_layer = profiler_layers[ndx_profiler]

        # Normalize optypes for comparison
        p_optype_norm = normalize_optype(p_layer['optype'])
        f_optype_norm = normalize_optype(f_layer['optype'])

        # Case 1: optypes match (after normalization)
        if p_optype_norm == f_optype_norm:
            # Compare shapes
            input_match, input_details = compare_tensor_shapes(
                p_layer.get('input_tensors', []),
                f_layer.get('input_tensors', []),
                strip_leading_ones,
                p_layer['optype'],
                strip_singleton_dims=strip_singleton_dims,
            )
            output_match, output_details = compare_tensor_shapes(
                p_layer.get('output_tensors', []),
                f_layer.get('output_tensors', []),
                strip_leading_ones,
                p_layer['optype'],
                strip_singleton_dims=strip_singleton_dims,
            )

            # Special handling for binary ops (add/mul/sub) if input counts
            # or shapes don't match — one side may use scalar/untracked operands
            p_canonical = normalize_polaris_optype(p_layer['optype'])
            if not input_match and to_comparison_group(p_canonical) == 'binary':
                bin_valid, bin_details = validate_binary_compatibility(
                    p_layer.get('input_tensors', []),
                    f_layer.get('input_tensors', []),
                    strip_leading_ones,
                    strip_singleton_dims=strip_singleton_dims,
                )
                if bin_valid:
                    input_match = True
                    input_details = bin_details

            # Special handling for reshape if standard comparison fails
            if p_canonical == 'reshape' and (not input_match or not output_match):
                reshape_valid, reshape_details = validate_reshape_compatibility(
                    p_layer.get('input_tensors', []),
                    p_layer.get('output_tensors', []),
                    f_layer.get('output_tensors', []),
                    strip_leading_ones,
                    strip_singleton_dims=strip_singleton_dims,
                )
                if reshape_valid:
                    # Accept reshape as valid - inputs may differ but outputs are compatible
                    input_match = True
                    output_match = True
                    output_details = reshape_details

            # Attribute comparison (dtype, layout, memory)
            attr_ok = True
            attr_details_parts = []
            if input_match and output_match and not ignore_attrs:
                in_attr_ok, in_attr_det = compare_tensor_attributes(p_layer, f_layer, 'input')
                out_attr_ok, out_attr_det = compare_tensor_attributes(p_layer, f_layer, 'output')
                if not in_attr_ok:
                    attr_ok = False
                    attr_details_parts.append(f"input attrs: {in_attr_det}")
                if not out_attr_ok:
                    attr_ok = False
                    attr_details_parts.append(f"output attrs: {out_attr_det}")

            if input_match and output_match and attr_ok:
                print(f"✓ [P:{p_layer['seqno']}] [F:{f_layer['seqno']}] {p_layer['optype']}  "
                      f"in: {format_shapes(p_layer.get('input_tensors', []))} | "
                      f"out: {format_shapes(p_layer.get('output_tensors', []))}")
                stats.total_matches += 1
            elif input_match and output_match and not attr_ok:
                print(f"✗ attr [P:{p_layer['seqno']}] [F:{f_layer['seqno']}] {p_layer['optype']}")
                for part in attr_details_parts:
                    print(f"  {part}")
                stats.attr_mismatches += 1
            else:
                print(f"✗ shape [P:{p_layer['seqno']}] [F:{f_layer['seqno']}] {p_layer['optype']}")
                if not input_match:
                    print(f"  input: polaris={format_shapes(p_layer.get('input_tensors', []))} "
                          f"profiler={format_shapes(f_layer.get('input_tensors', []))} ({input_details})")
                    stats.input_shape_mismatches += 1
                if not output_match:
                    print(f"  output: polaris={format_shapes(p_layer.get('output_tensors', []))} "
                          f"profiler={format_shapes(f_layer.get('output_tensors', []))} ({output_details})")
                    stats.output_shape_mismatches += 1
                stats.shape_mismatches += 1

            ndx_polaris += 1
            ndx_profiler += 1
            continue

        # Case 2: optypes don't match - search forward in profiler only (polaris is pivot)
        profiler_match_idx = find_next_match(
            profiler_layers, ndx_profiler + 1, p_optype_norm, max_search_distance
        )

        # If found in profiler, skip profiler entries to get there
        if profiler_match_idx is not None:
            for i in range(ndx_profiler, profiler_match_idx):
                layer = profiler_layers[i]
                print(f"⊘ [F:{layer['seqno']}] {layer['optype']} (skipped in polaris)")
                stats.unmatched_profiler += 1

            # Move profiler to matched position, polaris stays to compare
            ndx_profiler = profiler_match_idx
            # Continue to compare at this position (will be handled in next iteration)
        else:
            # Polaris entry not found in profiler - mark and advance polaris only
            print(f"✗ name [P:{p_layer['seqno']}] --- {p_layer['optype']} (not in profiler)")
            stats.name_mismatches += 1
            ndx_polaris += 1

    return stats


def _signature_string_for_layer(
    layer: Dict[str, Any],
    strip_leading_ones: bool,
    strip_singleton_dims: bool,
) -> str:
    """Normalized in/out shape string for grouping (matches compare_layers strip rules)."""

    def fmt_slot(shapes: List[str]) -> str:
        parts: List[str] = []
        for sh in shapes or []:
            dims = parse_shape(sh)
            nd = normalize_shape(dims, strip_leading_ones, strip_singleton_dims)
            parts.append("x".join(str(x) for x in nd) if nd else "")
        return ";".join(parts)

    ins = fmt_slot(layer.get("input_tensors") or [])
    outs = fmt_slot(layer.get("output_tensors") or [])
    return f"in[{ins}] out[{outs}]"


def _print_signature_summary(
    layers: List[Dict[str, Any]],
    label: str,
    strip_leading_ones: bool,
    strip_singleton_dims: bool,
    *,
    include_perf: bool,
) -> None:
    """Rollup: count (and optionally ms / LUT) per (optype, shape signature)."""
    counts: DefaultDict[Tuple[str, str], int] = defaultdict(int)
    ms_totals: DefaultDict[Tuple[str, str], float] = defaultdict(float)
    lut_totals: DefaultDict[Tuple[str, str], int] = defaultdict(int)

    for layer in layers:
        optype = str(layer.get("optype", ""))
        sig = _signature_string_for_layer(layer, strip_leading_ones, strip_singleton_dims)
        key = (optype, sig)
        counts[key] += 1
        if include_perf:
            d = layer.get("duration_ms")
            if d is not None:
                ms_totals[key] += float(d)
            if layer.get("uses_perf_lookup"):
                lut_totals[key] += 1

    keys_sorted = sorted(counts.keys(), key=lambda k: (-counts[k], k[0], k[1]))
    any_lut = include_perf and sum(lut_totals.values()) > 0

    print(f"\n{'=' * 72}")
    print(f"  Summary by layer type + signature ({label})")
    print(f"{'=' * 72}")

    hdr_count = "Count"
    hdr_type = "Layer type"
    hdr_sig = "Signature (normalized in / out)"
    col_c = max(5, len(hdr_count))
    col_t = max(12, max((len(k[0]) for k in keys_sorted), default=len(hdr_type)))
    col_m = 11
    col_l = 9

    if include_perf:
        hdr_ms = "Sum ms"
        hdr_lut = "LUT"
        print(
            f"  {hdr_count:>{col_c}}  {hdr_type:<{col_t}}  {hdr_ms:>{col_m}}  "
            f"{hdr_lut:>{col_l}}  {hdr_sig}"
        )
    else:
        print(f"  {hdr_count:>{col_c}}  {hdr_type:<{col_t}}  {hdr_sig}")
    print(f"  {'─' * 72}")

    total_n = 0
    total_ms = 0.0
    total_lut = 0
    for key in keys_sorted:
        n = counts[key]
        total_n += n
        op, sig = key
        if include_perf:
            ms = ms_totals.get(key, 0.0)
            lut = lut_totals.get(key, 0)
            total_ms += ms
            total_lut += lut
            lut_s = f"{lut}/{n}" if any_lut else "—"
            print(
                f"  {n:>{col_c}}  {op:<{col_t}}  {ms:>{col_m}.4f}  {lut_s:>{col_l}}  {sig}"
            )
        else:
            print(f"  {n:>{col_c}}  {op:<{col_t}}  {sig}")

    print(f"  {'─' * 72}")
    if include_perf:
        lut_footer = f"{total_lut}/{total_n}" if any_lut else "—"
        print(
            f"  {total_n:>{col_c}}  {'TOTAL':<{col_t}}  {total_ms:>{col_m}.4f}  {lut_footer:>{col_l}}"
        )
    else:
        print(f"  {total_n:>{col_c}}  {'TOTAL':<{col_t}}")
    print()


def print_summary(stats: ComparisonStats) -> None:
    """Print summary statistics."""
    print("\n=== Summary ===")
    print(f"Total matches: {stats.total_matches}")
    print(f"Name mismatches: {stats.name_mismatches}")
    print(f"Shape mismatches: {stats.shape_mismatches} "
          f"({stats.input_shape_mismatches} input, {stats.output_shape_mismatches} output)")
    print(f"Attribute mismatches: {stats.attr_mismatches}")
    print(f"Unmatched entries: {stats.unmatched_polaris + stats.unmatched_profiler} "
          f"({stats.unmatched_polaris} polaris, {stats.unmatched_profiler} profiler)")
    print(f"Ambiguous: {stats.ambiguous}")


# ---------------------------------------------------------------------------
# Performance summary helpers
# ---------------------------------------------------------------------------

def _aggregate_duration_by_optype_signature(
    layers: List[Dict[str, Any]],
    strip_leading_ones: bool,
    strip_singleton_dims: bool,
) -> Dict[Tuple[str, str], Tuple[int, float, int]]:
    """Group layers by (optype, normalized shape signature); sum durations and LUT hits."""
    totals: DefaultDict[Tuple[str, str], float] = defaultdict(float)
    counts: DefaultDict[Tuple[str, str], int] = defaultdict(int)
    lut_hits: DefaultDict[Tuple[str, str], int] = defaultdict(int)
    for layer in layers:
        optype = str(layer.get("optype", ""))
        sig = _signature_string_for_layer(layer, strip_leading_ones, strip_singleton_dims)
        key = (optype, sig)
        counts[key] += 1
        dur = layer.get("duration_ms")
        if dur is not None:
            totals[key] += float(dur)
        if layer.get("uses_perf_lookup"):
            lut_hits[key] += 1
    all_keys = set(counts) | set(totals) | set(lut_hits)
    return {
        k: (counts[k], totals.get(k, 0.0), lut_hits.get(k, 0))
        for k in all_keys
    }


def _aggregate_duration_by_optype(
    layers: List[Dict[str, Any]],
) -> Dict[str, Tuple[int, float, int]]:
    """Group layers by fine-grained canonical optype and sum durations.

    Returns ``{optype: (count, total_ms, lut_hits)}`` sorted by descending
    total_ms.  Layers whose ``duration_ms`` is None are counted but
    contribute 0 ms.  ``lut_hits`` counts layers where
    ``uses_perf_lookup`` is True (Polaris-only; always 0 for profiler).
    """
    totals: Dict[str, float] = defaultdict(float)
    counts: Dict[str, int] = defaultdict(int)
    lut_hits: Dict[str, int] = defaultdict(int)
    for layer in layers:
        optype = layer['optype']
        counts[optype] += 1
        dur = layer.get('duration_ms')
        if dur is not None:
            totals[optype] += dur
        if layer.get('uses_perf_lookup'):
            lut_hits[optype] += 1
    all_optypes = set(counts) | set(totals)
    return {
        op: (counts[op], totals.get(op, 0.0), lut_hits.get(op, 0))
        for op in sorted(all_optypes, key=lambda o: totals.get(o, 0.0), reverse=True)
    }


def _pct_gap(reference: float, other: float) -> str:
    """Format percentage gap of *other* w.r.t. *reference*.

    Positive means *other* is larger (slower).
    """
    if reference == 0.0:
        return "N/A"
    gap = (other - reference) / reference * 100.0
    sign = "+" if gap >= 0 else ""
    return f"{sign}{gap:.2f}%"


def _print_perf_standalone_by_signature(
    layers: List[Dict[str, Any]],
    source_label: str,
    strip_leading_ones: bool,
    strip_singleton_dims: bool,
) -> None:
    """Standalone performance breakdown grouped by optype + signature."""
    by_key = _aggregate_duration_by_optype_signature(
        layers, strip_leading_ones, strip_singleton_dims
    )
    total_ms = sum(ms for _, ms, _ in by_key.values())
    total_count = sum(cnt for cnt, _, _ in by_key.values())
    total_lut = sum(lut for _, _, lut in by_key.values())
    has_lut = total_lut > 0

    print(f"\n{'=' * 60}")
    print(f"  Performance Summary by type + signature ({source_label})")
    print(f"{'=' * 60}")
    print(f"\n  Network total: {total_ms:.4f} ms")
    if has_lut:
        print(f"  LUT hits: {total_lut}/{total_count}")
    print()

    keys_sorted = sorted(
        by_key.keys(),
        key=lambda k: by_key.get(k, (0, 0.0, 0))[1],
        reverse=True,
    )

    hdr_type = "Layer type"
    hdr_sig = "Signature"
    col_w_t = max(10, max((len(k[0]) for k in keys_sorted), default=len(hdr_type)))
    col_w_s = max(24, min(72, max((len(k[1]) for k in keys_sorted), default=len(hdr_sig))))

    hdr_cnt = "Count"
    hdr_dur = "Duration (ms)"
    hdr_lut = "LUT"
    col_w_cnt = max(len(hdr_cnt), 6)
    col_w_dur = max(len(hdr_dur), 14)
    col_w_lut = max(len(hdr_lut), 8)

    header = (
        f"  {hdr_type:<{col_w_t}}  {hdr_sig:<{col_w_s}}  {hdr_cnt:>{col_w_cnt}}  "
        f"{hdr_dur:>{col_w_dur}}"
    )
    if has_lut:
        header += f"  {hdr_lut:>{col_w_lut}}"
    print(header)
    rule_len = col_w_t + col_w_s + col_w_cnt + col_w_dur + 8 + (col_w_lut + 2 if has_lut else 0)
    print(f"  {'─' * min(rule_len, 120)}")

    for key in keys_sorted:
        op, sig = key
        cnt, ms, lut = by_key[key]
        sig_disp = sig if len(sig) <= col_w_s else sig[: col_w_s - 3] + "..."
        line = (
            f"  {op:<{col_w_t}}  {sig_disp:<{col_w_s}}  {cnt:>{col_w_cnt}}  "
            f"{ms:>{col_w_dur}.4f}"
        )
        if has_lut:
            line += f"  {f'{lut}/{cnt}':>{col_w_lut}}"
        print(line)

    print(f"  {'─' * min(rule_len, 120)}")
    line = (
        f"  {'TOTAL':<{col_w_t}}  {'':<{col_w_s}}  {total_count:>{col_w_cnt}}  "
        f"{total_ms:>{col_w_dur}.4f}"
    )
    if has_lut:
        line += f"  {f'{total_lut}/{total_count}':>{col_w_lut}}"
    print(line)
    print()


def _print_perf_standalone(
    layers: List[Dict[str, Any]],
    source_label: str,
) -> None:
    """Print a standalone performance breakdown for a single source."""
    by_optype = _aggregate_duration_by_optype(layers)
    total_ms = sum(ms for _, ms, _ in by_optype.values())
    total_count = sum(cnt for cnt, _, _ in by_optype.values())
    total_lut = sum(lut for _, _, lut in by_optype.values())
    has_lut = total_lut > 0

    print(f"\n{'=' * 60}")
    print(f"  Performance Summary ({source_label})")
    print(f"{'=' * 60}")
    print(f"\n  Network total: {total_ms:.4f} ms")
    if has_lut:
        print(f"  LUT hits: {total_lut}/{total_count}")
    print()

    hdr_type = "Layer Type"
    hdr_cnt = "Count"
    hdr_dur = "Duration (ms)"
    hdr_lut = "LUT"
    col_w_type = max(len(hdr_type), max((len(op) for op in by_optype), default=10))
    col_w_cnt = max(len(hdr_cnt), 6)
    col_w_dur = max(len(hdr_dur), 14)
    col_w_lut = max(len(hdr_lut), 8)

    header = f"  {hdr_type:<{col_w_type}}  {hdr_cnt:>{col_w_cnt}}  {hdr_dur:>{col_w_dur}}"
    if has_lut:
        header += f"  {hdr_lut:>{col_w_lut}}"
    print(header)
    rule_len = col_w_type + col_w_cnt + col_w_dur + 4 + (col_w_lut + 2 if has_lut else 0)
    print(f"  {'─' * rule_len}")

    for op, (cnt, ms, lut) in by_optype.items():
        line = f"  {op:<{col_w_type}}  {cnt:>{col_w_cnt}}  {ms:>{col_w_dur}.4f}"
        if has_lut:
            line += f"  {f'{lut}/{cnt}':>{col_w_lut}}"
        print(line)

    print(f"  {'─' * rule_len}")
    line = f"  {'TOTAL':<{col_w_type}}  {total_count:>{col_w_cnt}}  {total_ms:>{col_w_dur}.4f}"
    if has_lut:
        line += f"  {f'{total_lut}/{total_count}':>{col_w_lut}}"
    print(line)
    print()


def _print_perf_comparison(
    profiler_layers: List[Dict[str, Any]],
    polaris_layers: List[Dict[str, Any]],
    *,
    by_signature: bool = False,
    strip_leading_ones: bool = False,
    strip_singleton_dims: bool = False,
) -> None:
    """Print side-by-side performance comparison with gap w.r.t. profiler."""
    prof_by: Union[Dict[Tuple[str, str], Tuple[int, float, int]], Dict[str, Tuple[int, float, int]]]
    pol_by: Union[Dict[Tuple[str, str], Tuple[int, float, int]], Dict[str, Tuple[int, float, int]]]

    if by_signature:
        prof_by = _aggregate_duration_by_optype_signature(
            profiler_layers, strip_leading_ones, strip_singleton_dims
        )
        pol_by = _aggregate_duration_by_optype_signature(
            polaris_layers, strip_leading_ones, strip_singleton_dims
        )
        title = "Performance Summary (by layer type + signature)"
    else:
        prof_by = _aggregate_duration_by_optype(profiler_layers)
        pol_by = _aggregate_duration_by_optype(polaris_layers)
        title = "Performance Summary"

    prof_total_ms = sum(ms for _, ms, _ in prof_by.values())
    pol_total_ms = sum(ms for _, ms, _ in pol_by.values())
    prof_total_cnt = sum(cnt for cnt, _, _ in prof_by.values())
    pol_total_cnt = sum(cnt for cnt, _, _ in pol_by.values())
    pol_total_lut = sum(lut for _, _, lut in pol_by.values())

    print(f"\n{'=' * 82}")
    print(f"  {title}")
    print(f"{'=' * 82}")

    print("\n  Network total:")
    print(f"    Profiler:  {prof_total_ms:.4f} ms")
    print(f"    Polaris:   {pol_total_ms:.4f} ms")
    print(f"    Gap:       {_pct_gap(prof_total_ms, pol_total_ms)} (w.r.t. profiler)")
    print(f"    Polaris LUT hits: {pol_total_lut}/{pol_total_cnt}")
    print()

    all_keys: List[Any] = list(dict.fromkeys(list(prof_by.keys()) + list(pol_by.keys())))
    all_keys.sort(key=lambda k: prof_by.get(k, (0, 0.0, 0))[1], reverse=True)

    if by_signature:
        col_w_op = max(10, max((len(k[0]) for k in all_keys), default=10))
        col_w_sig = max(28, min(56, max((len(k[1]) for k in all_keys), default=28)))
        hdr = (
            f"  {'Layer type':<{col_w_op}}"
            f"  {'Signature':<{col_w_sig}}"
            f"  {'#Prof':>6}  {'Profiler(ms)':>13}"
            f"  {'#Pol':>6}  {'Polaris(ms)':>13}"
            f"  {'LUT':>8}"
            f"  {'Abs Gap(ms)':>12}"
            f"  {'Gap%':>9}"
        )
    else:
        col_w_op = max(10, max((len(str(op)) for op in all_keys), default=10))
        col_w_sig = 0
        hdr = (
            f"  {'Layer Type':<{col_w_op}}"
            f"  {'#Prof':>6}  {'Profiler(ms)':>13}"
            f"  {'#Pol':>6}  {'Polaris(ms)':>13}"
            f"  {'LUT':>8}"
            f"  {'Abs Gap(ms)':>12}"
            f"  {'Gap%':>9}"
        )
    print(hdr)
    rule_len = col_w_op + 6 + 13 + 6 + 13 + 8 + 12 + 9 + 14 + (col_w_sig + 2 if by_signature else 0)
    print(f"  {'─' * min(rule_len, 120)}")

    for key in all_keys:
        if by_signature:
            op, sig = key
            sig_disp = sig if len(sig) <= col_w_sig else sig[: col_w_sig - 3] + "..."
            p_cnt, p_ms, _ = prof_by.get(key, (0, 0.0, 0))
            s_cnt, s_ms, s_lut = pol_by.get(key, (0, 0.0, 0))
        else:
            op = key
            sig_disp = ""
            p_cnt, p_ms, _ = prof_by.get(op, (0, 0.0, 0))
            s_cnt, s_ms, s_lut = pol_by.get(op, (0, 0.0, 0))
        gap_pct = _pct_gap(p_ms, s_ms)
        abs_gap = s_ms - p_ms
        abs_gap_s = f"{abs_gap:+.4f}" if (p_cnt and s_cnt) else "—"
        p_cnt_s = str(p_cnt) if p_cnt else "—"
        s_cnt_s = str(s_cnt) if s_cnt else "—"
        p_ms_s = f"{p_ms:.4f}" if p_cnt else "—"
        s_ms_s = f"{s_ms:.4f}" if s_cnt else "—"
        lut_s = f"{s_lut}/{s_cnt}" if s_cnt else "—"
        if by_signature:
            print(
                f"  {op:<{col_w_op}}"
                f"  {sig_disp:<{col_w_sig}}"
                f"  {p_cnt_s:>6}  {p_ms_s:>13}"
                f"  {s_cnt_s:>6}  {s_ms_s:>13}"
                f"  {lut_s:>8}"
                f"  {abs_gap_s:>12}"
                f"  {gap_pct:>9}"
            )
        else:
            print(
                f"  {op:<{col_w_op}}"
                f"  {p_cnt_s:>6}  {p_ms_s:>13}"
                f"  {s_cnt_s:>6}  {s_ms_s:>13}"
                f"  {lut_s:>8}"
                f"  {abs_gap_s:>12}"
                f"  {gap_pct:>9}"
            )

    abs_total = pol_total_ms - prof_total_ms
    print(f"  {'─' * min(rule_len, 120)}")
    if by_signature:
        print(
            f"  {'TOTAL':<{col_w_op}}"
            f"  {'':<{col_w_sig}}"
            f"  {prof_total_cnt:>6}  {prof_total_ms:>13.4f}"
            f"  {pol_total_cnt:>6}  {pol_total_ms:>13.4f}"
            f"  {f'{pol_total_lut}/{pol_total_cnt}':>8}"
            f"  {abs_total:>+12.4f}"
            f"  {_pct_gap(prof_total_ms, pol_total_ms):>9}"
        )
    else:
        print(
            f"  {'TOTAL':<{col_w_op}}"
            f"  {prof_total_cnt:>6}  {prof_total_ms:>13.4f}"
            f"  {pol_total_cnt:>6}  {pol_total_ms:>13.4f}"
            f"  {f'{pol_total_lut}/{pol_total_cnt}':>8}"
            f"  {abs_total:>+12.4f}"
            f"  {_pct_gap(prof_total_ms, pol_total_ms):>9}"
        )
    print()


# ---------------------------------------------------------------------------
# XLSX report (three sheets: Summary, By Layer Type, By Layer Signature)
# ---------------------------------------------------------------------------

def _write_xlsx_report(
    path: str,
    *,
    polaris_layers: Optional[List[Dict[str, Any]]],
    profiler_layers: Optional[List[Dict[str, Any]]],
    stats: Optional[ComparisonStats],
    polaris_label: str,
    profiler_label: str,
    polaris_source: Optional[str],
    profiler_source: Optional[str],
    strip_leading_ones: bool,
    strip_singleton_dims: bool,
) -> None:
    """Write a 3-sheet XLSX comparison report.

    Sheets:
      1. "Summary"          — model-wide totals, comparison, shape/attr stats.
      2. "By Layer Type"    — per canonical optype rollup with comparison.
      3. "By Layer Signature" — per (optype + normalized shape signature).

    In two-file mode (both ``polaris_layers`` and ``profiler_layers`` provided)
    each sheet includes Profiler vs Polaris columns and the absolute / percent
    gap.  In single-file mode only the side that was loaded is emitted.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:  # pragma: no cover - import guarded at call site
        raise RuntimeError(
            "openpyxl is required for --xlsx output. Install it via "
            "`pip install openpyxl` (already present in the polarisdev env)."
        ) from e

    have_polaris = polaris_layers is not None
    have_profiler = profiler_layers is not None
    two_sided = have_polaris and have_profiler

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    total_font = Font(bold=True, italic=True)
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _style_header(ws, row_idx: int, n_cols: int) -> None:
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

    def _style_total(ws, row_idx: int, n_cols: int) -> None:
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = total_font
            cell.fill = total_fill

    def _autosize(ws, n_cols: int, max_width: int = 60) -> None:
        for c in range(1, n_cols + 1):
            letter = get_column_letter(c)
            best = 8
            for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
                v = row[0]
                if v is None:
                    continue
                if isinstance(v, float):
                    s = f"{v:.4f}"
                else:
                    s = str(v)
                if len(s) > best:
                    best = len(s)
            ws.column_dimensions[letter].width = min(best + 2, max_width)

    def _pct(reference: float, other: float) -> Optional[float]:
        if reference == 0.0:
            return None
        return (other - reference) / reference * 100.0

    wb = Workbook()

    # ------------------ Sheet 1: Summary ------------------
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Summary"

    ws1.cell(row=1, column=1, value="Compare Layers — XLSX Report").font = Font(bold=True, size=13)
    r = 2
    if profiler_source:
        ws1.cell(row=r, column=1, value="Profiler CSV"); ws1.cell(row=r, column=2, value=profiler_source); r += 1
    if polaris_source:
        ws1.cell(row=r, column=1, value="Polaris CSV"); ws1.cell(row=r, column=2, value=polaris_source); r += 1
    ws1.cell(row=r, column=1, value="strip_leading_ones"); ws1.cell(row=r, column=2, value=bool(strip_leading_ones)); r += 1
    ws1.cell(row=r, column=1, value="strip_singleton_dims"); ws1.cell(row=r, column=2, value=bool(strip_singleton_dims)); r += 1
    r += 1

    # Network totals
    ws1.cell(row=r, column=1, value="Network totals").font = Font(bold=True); r += 1
    if two_sided:
        headers = ["Metric", profiler_label, polaris_label, "Abs Gap", "Gap %"]
    elif have_profiler:
        headers = ["Metric", profiler_label]
    else:
        headers = ["Metric", polaris_label]
    for c, h in enumerate(headers, start=1):
        ws1.cell(row=r, column=c, value=h)
    _style_header(ws1, r, len(headers))
    r += 1

    def _layer_totals(layers: List[Dict[str, Any]]) -> Tuple[int, float, int]:
        cnt = len(layers)
        ms = sum((float(l["duration_ms"]) for l in layers if l.get("duration_ms") is not None), 0.0)
        lut = sum(1 for l in layers if l.get("uses_perf_lookup"))
        return cnt, ms, lut

    prof_cnt = prof_ms = prof_lut = 0
    pol_cnt = pol_ms = pol_lut = 0
    if have_profiler:
        prof_cnt, prof_ms, prof_lut = _layer_totals(profiler_layers)  # type: ignore[assignment,arg-type]
    if have_polaris:
        pol_cnt, pol_ms, pol_lut = _layer_totals(polaris_layers)  # type: ignore[assignment,arg-type]

    def _write_metric(metric: str, prof_v, pol_v, *, fmt_ms: bool = False, percent: bool = False) -> None:
        nonlocal r
        ws1.cell(row=r, column=1, value=metric)
        if two_sided:
            ws1.cell(row=r, column=2, value=prof_v)
            ws1.cell(row=r, column=3, value=pol_v)
            if isinstance(prof_v, (int, float)) and isinstance(pol_v, (int, float)):
                ws1.cell(row=r, column=4, value=pol_v - prof_v)
                p = _pct(float(prof_v), float(pol_v))
                ws1.cell(row=r, column=5, value=(p if p is not None else "N/A"))
                if fmt_ms:
                    ws1.cell(row=r, column=2).number_format = "0.0000"
                    ws1.cell(row=r, column=3).number_format = "0.0000"
                    ws1.cell(row=r, column=4).number_format = "+0.0000;-0.0000"
                if p is not None:
                    ws1.cell(row=r, column=5).number_format = "+0.00\"%\";-0.00\"%\""
        elif have_profiler:
            ws1.cell(row=r, column=2, value=prof_v)
            if fmt_ms and isinstance(prof_v, (int, float)):
                ws1.cell(row=r, column=2).number_format = "0.0000"
        else:
            ws1.cell(row=r, column=2, value=pol_v)
            if fmt_ms and isinstance(pol_v, (int, float)):
                ws1.cell(row=r, column=2).number_format = "0.0000"
        r += 1

    _write_metric("Total layers", prof_cnt, pol_cnt)
    _write_metric("Total duration (ms)", prof_ms, pol_ms, fmt_ms=True)
    if have_polaris:
        # Polaris-only metric — display only on Polaris column when single-sided.
        pol_miss = pol_cnt - pol_lut
        if two_sided:
            ws1.cell(row=r, column=1, value="Polaris LUT hits (count / total)")
            ws1.cell(row=r, column=2, value="—")
            ws1.cell(row=r, column=3, value=f"{pol_lut} / {pol_cnt}")
            r += 1
            ws1.cell(row=r, column=1, value="Polaris LUT misses (count / total)")
            ws1.cell(row=r, column=2, value="—")
            ws1.cell(row=r, column=3, value=f"{pol_miss} / {pol_cnt}")
            r += 1
        else:
            ws1.cell(row=r, column=1, value="Polaris LUT hits (count / total)")
            ws1.cell(row=r, column=2, value=f"{pol_lut} / {pol_cnt}")
            r += 1
            ws1.cell(row=r, column=1, value="Polaris LUT misses (count / total)")
            ws1.cell(row=r, column=2, value=f"{pol_miss} / {pol_cnt}")
            r += 1

    # Shape/attr stats (two-file only)
    if two_sided and stats is not None:
        r += 1
        ws1.cell(row=r, column=1, value="Shape / attribute comparison").font = Font(bold=True); r += 1
        for c, h in enumerate(["Metric", "Count"], start=1):
            ws1.cell(row=r, column=c, value=h)
        _style_header(ws1, r, 2); r += 1
        for label, val in [
            ("Total matches", stats.total_matches),
            ("Name mismatches", stats.name_mismatches),
            ("Shape mismatches", stats.shape_mismatches),
            ("  input shape mismatches", stats.input_shape_mismatches),
            ("  output shape mismatches", stats.output_shape_mismatches),
            ("Attribute mismatches", stats.attr_mismatches),
            ("Unmatched (Polaris)", stats.unmatched_polaris),
            ("Unmatched (Profiler)", stats.unmatched_profiler),
            ("Ambiguous", stats.ambiguous),
        ]:
            ws1.cell(row=r, column=1, value=label)
            ws1.cell(row=r, column=2, value=val)
            r += 1

    _autosize(ws1, 5 if two_sided else 2)

    # ---------- Sheet 2: By Layer Type ----------
    ws2 = wb.create_sheet("By Layer Type")
    prof_by_op: Dict[str, Tuple[int, float, int]] = (
        _aggregate_duration_by_optype(profiler_layers) if have_profiler else {}  # type: ignore[arg-type]
    )
    pol_by_op: Dict[str, Tuple[int, float, int]] = (
        _aggregate_duration_by_optype(polaris_layers) if have_polaris else {}  # type: ignore[arg-type]
    )
    keys_op = list(dict.fromkeys(list(prof_by_op.keys()) + list(pol_by_op.keys())))
    keys_op.sort(
        key=lambda k: max(prof_by_op.get(k, (0, 0.0, 0))[1], pol_by_op.get(k, (0, 0.0, 0))[1]),
        reverse=True,
    )
    if two_sided:
        op_headers = [
            "Layer Type",
            f"# {profiler_label}", f"{profiler_label} ms",
            f"# {polaris_label}", f"{polaris_label} ms",
            "Polaris LUT hit", "Polaris LUT miss", "Polaris LUT total",
            "Only in Polaris", "Only in Hardware",
            "Abs Gap (ms)", "Gap %",
        ]
    elif have_profiler:
        op_headers = ["Layer Type", f"# {profiler_label}", f"{profiler_label} ms"]
    else:
        op_headers = [
            "Layer Type", f"# {polaris_label}", f"{polaris_label} ms",
            "Polaris LUT hit", "Polaris LUT miss", "Polaris LUT total",
        ]
    for c, h in enumerate(op_headers, start=1):
        ws2.cell(row=1, column=c, value=h)
    _style_header(ws2, 1, len(op_headers))

    rr = 2
    only_pol_op_total = 0
    only_hw_op_total = 0
    for k in keys_op:
        p_cnt, p_ms, _ = prof_by_op.get(k, (0, 0.0, 0))
        s_cnt, s_ms, s_lut = pol_by_op.get(k, (0, 0.0, 0))
        s_miss = s_cnt - s_lut
        only_pol = max(0, s_cnt - p_cnt)
        only_hw = max(0, p_cnt - s_cnt)
        only_pol_op_total += only_pol
        only_hw_op_total += only_hw
        if two_sided:
            ws2.cell(row=rr, column=1, value=k)
            ws2.cell(row=rr, column=2, value=p_cnt)
            ws2.cell(row=rr, column=3, value=p_ms).number_format = "0.0000"
            ws2.cell(row=rr, column=4, value=s_cnt)
            ws2.cell(row=rr, column=5, value=s_ms).number_format = "0.0000"
            ws2.cell(row=rr, column=6, value=s_lut)
            ws2.cell(row=rr, column=7, value=s_miss)
            ws2.cell(row=rr, column=8, value=s_cnt)
            ws2.cell(row=rr, column=9, value=only_pol)
            ws2.cell(row=rr, column=10, value=only_hw)
            ws2.cell(row=rr, column=11, value=(s_ms - p_ms)).number_format = "+0.0000;-0.0000"
            p = _pct(p_ms, s_ms)
            cell = ws2.cell(row=rr, column=12, value=(p if p is not None else "N/A"))
            if p is not None:
                cell.number_format = "+0.00\"%\";-0.00\"%\""
        elif have_profiler:
            ws2.cell(row=rr, column=1, value=k)
            ws2.cell(row=rr, column=2, value=p_cnt)
            ws2.cell(row=rr, column=3, value=p_ms).number_format = "0.0000"
        else:
            ws2.cell(row=rr, column=1, value=k)
            ws2.cell(row=rr, column=2, value=s_cnt)
            ws2.cell(row=rr, column=3, value=s_ms).number_format = "0.0000"
            ws2.cell(row=rr, column=4, value=s_lut)
            ws2.cell(row=rr, column=5, value=s_miss)
            ws2.cell(row=rr, column=6, value=s_cnt)
        rr += 1

    # TOTAL row
    pol_miss_total = pol_cnt - pol_lut
    ws2.cell(row=rr, column=1, value="TOTAL")
    if two_sided:
        ws2.cell(row=rr, column=2, value=prof_cnt)
        ws2.cell(row=rr, column=3, value=prof_ms).number_format = "0.0000"
        ws2.cell(row=rr, column=4, value=pol_cnt)
        ws2.cell(row=rr, column=5, value=pol_ms).number_format = "0.0000"
        ws2.cell(row=rr, column=6, value=pol_lut)
        ws2.cell(row=rr, column=7, value=pol_miss_total)
        ws2.cell(row=rr, column=8, value=pol_cnt)
        ws2.cell(row=rr, column=9, value=only_pol_op_total)
        ws2.cell(row=rr, column=10, value=only_hw_op_total)
        ws2.cell(row=rr, column=11, value=(pol_ms - prof_ms)).number_format = "+0.0000;-0.0000"
        p = _pct(prof_ms, pol_ms)
        cell = ws2.cell(row=rr, column=12, value=(p if p is not None else "N/A"))
        if p is not None:
            cell.number_format = "+0.00\"%\";-0.00\"%\""
    elif have_profiler:
        ws2.cell(row=rr, column=2, value=prof_cnt)
        ws2.cell(row=rr, column=3, value=prof_ms).number_format = "0.0000"
    else:
        ws2.cell(row=rr, column=2, value=pol_cnt)
        ws2.cell(row=rr, column=3, value=pol_ms).number_format = "0.0000"
        ws2.cell(row=rr, column=4, value=pol_lut)
        ws2.cell(row=rr, column=5, value=pol_miss_total)
        ws2.cell(row=rr, column=6, value=pol_cnt)
    _style_total(ws2, rr, len(op_headers))
    ws2.freeze_panes = "A2"
    _autosize(ws2, len(op_headers))

    # ---------- Sheet 3: By Layer Signature ----------
    ws3 = wb.create_sheet("By Layer Signature")
    prof_by_sig: Dict[Tuple[str, str], Tuple[int, float, int]] = (
        _aggregate_duration_by_optype_signature(
            profiler_layers, strip_leading_ones, strip_singleton_dims  # type: ignore[arg-type]
        ) if have_profiler else {}
    )
    pol_by_sig: Dict[Tuple[str, str], Tuple[int, float, int]] = (
        _aggregate_duration_by_optype_signature(
            polaris_layers, strip_leading_ones, strip_singleton_dims  # type: ignore[arg-type]
        ) if have_polaris else {}
    )
    keys_sig = list(dict.fromkeys(list(prof_by_sig.keys()) + list(pol_by_sig.keys())))
    keys_sig.sort(
        key=lambda k: max(
            prof_by_sig.get(k, (0, 0.0, 0))[1], pol_by_sig.get(k, (0, 0.0, 0))[1]
        ),
        reverse=True,
    )
    if two_sided:
        sig_headers = [
            "Layer Type", "Signature",
            f"# {profiler_label}", f"{profiler_label} ms",
            f"# {polaris_label}", f"{polaris_label} ms",
            "Polaris LUT hit", "Polaris LUT miss", "Polaris LUT total",
            "Only in Polaris", "Only in Hardware",
            "Abs Gap (ms)", "Gap %",
        ]
    elif have_profiler:
        sig_headers = ["Layer Type", "Signature", f"# {profiler_label}", f"{profiler_label} ms"]
    else:
        sig_headers = [
            "Layer Type", "Signature",
            f"# {polaris_label}", f"{polaris_label} ms",
            "Polaris LUT hit", "Polaris LUT miss", "Polaris LUT total",
        ]
    for c, h in enumerate(sig_headers, start=1):
        ws3.cell(row=1, column=c, value=h)
    _style_header(ws3, 1, len(sig_headers))

    rr = 2
    only_pol_sig_total = 0
    only_hw_sig_total = 0
    for sig_key in keys_sig:
        op, sig = sig_key
        p_cnt, p_ms, _ = prof_by_sig.get(sig_key, (0, 0.0, 0))
        s_cnt, s_ms, s_lut = pol_by_sig.get(sig_key, (0, 0.0, 0))
        s_miss = s_cnt - s_lut
        only_pol = max(0, s_cnt - p_cnt)
        only_hw = max(0, p_cnt - s_cnt)
        only_pol_sig_total += only_pol
        only_hw_sig_total += only_hw
        ws3.cell(row=rr, column=1, value=op)
        ws3.cell(row=rr, column=2, value=sig).alignment = left
        if two_sided:
            ws3.cell(row=rr, column=3, value=p_cnt)
            ws3.cell(row=rr, column=4, value=p_ms).number_format = "0.0000"
            ws3.cell(row=rr, column=5, value=s_cnt)
            ws3.cell(row=rr, column=6, value=s_ms).number_format = "0.0000"
            ws3.cell(row=rr, column=7, value=s_lut)
            ws3.cell(row=rr, column=8, value=s_miss)
            ws3.cell(row=rr, column=9, value=s_cnt)
            ws3.cell(row=rr, column=10, value=only_pol)
            ws3.cell(row=rr, column=11, value=only_hw)
            ws3.cell(row=rr, column=12, value=(s_ms - p_ms)).number_format = "+0.0000;-0.0000"
            p = _pct(p_ms, s_ms)
            cell = ws3.cell(row=rr, column=13, value=(p if p is not None else "N/A"))
            if p is not None:
                cell.number_format = "+0.00\"%\";-0.00\"%\""
        elif have_profiler:
            ws3.cell(row=rr, column=3, value=p_cnt)
            ws3.cell(row=rr, column=4, value=p_ms).number_format = "0.0000"
        else:
            ws3.cell(row=rr, column=3, value=s_cnt)
            ws3.cell(row=rr, column=4, value=s_ms).number_format = "0.0000"
            ws3.cell(row=rr, column=5, value=s_lut)
            ws3.cell(row=rr, column=6, value=s_miss)
            ws3.cell(row=rr, column=7, value=s_cnt)
        rr += 1

    pol_miss_total = pol_cnt - pol_lut
    ws3.cell(row=rr, column=1, value="TOTAL")
    ws3.cell(row=rr, column=2, value="")
    if two_sided:
        ws3.cell(row=rr, column=3, value=prof_cnt)
        ws3.cell(row=rr, column=4, value=prof_ms).number_format = "0.0000"
        ws3.cell(row=rr, column=5, value=pol_cnt)
        ws3.cell(row=rr, column=6, value=pol_ms).number_format = "0.0000"
        ws3.cell(row=rr, column=7, value=pol_lut)
        ws3.cell(row=rr, column=8, value=pol_miss_total)
        ws3.cell(row=rr, column=9, value=pol_cnt)
        ws3.cell(row=rr, column=10, value=only_pol_sig_total)
        ws3.cell(row=rr, column=11, value=only_hw_sig_total)
        ws3.cell(row=rr, column=12, value=(pol_ms - prof_ms)).number_format = "+0.0000;-0.0000"
        p = _pct(prof_ms, pol_ms)
        cell = ws3.cell(row=rr, column=13, value=(p if p is not None else "N/A"))
        if p is not None:
            cell.number_format = "+0.00\"%\";-0.00\"%\""
    elif have_profiler:
        ws3.cell(row=rr, column=3, value=prof_cnt)
        ws3.cell(row=rr, column=4, value=prof_ms).number_format = "0.0000"
    else:
        ws3.cell(row=rr, column=3, value=pol_cnt)
        ws3.cell(row=rr, column=4, value=pol_ms).number_format = "0.0000"
        ws3.cell(row=rr, column=5, value=pol_lut)
        ws3.cell(row=rr, column=6, value=pol_miss_total)
        ws3.cell(row=rr, column=7, value=pol_cnt)
    _style_total(ws3, rr, len(sig_headers))
    ws3.freeze_panes = "C2"
    _autosize(ws3, len(sig_headers))

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    """Main entry point."""
    args = parse_args()

    # Sanitize and validate file paths to prevent path traversal attacks
    try:
        file1_path = sanitize_file_path(args.file1)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    file2_path = None
    if args.file2 is not None:
        try:
            file2_path = sanitize_file_path(args.file2)
        except ValueError as e:
            print(f"Error: {e}", file=sys.stderr)
            return 1

    # --- Standalone mode (single file + --perf and/or --summarize-by-signature) ---
    if file2_path is None:
        if not args.perf and not args.summarize_by_signature:
            print("Error: Two files are required for shape comparison. "
                  "Use --perf and/or --summarize-by-signature with a single file.",
                  file=sys.stderr)
            return 1

        ftype = detect_file_type(str(file1_path))
        if ftype is None:
            print("Error: Could not determine file type. Expected CSV with "
                  "'archname' (polaris) or 'OP CODE' (profiler) columns.",
                  file=sys.stderr)
            return 1

        try:
            if ftype == 'polaris':
                layers = layers_polaris(str(file1_path))
                label = "Polaris"
            else:
                layers = layers_profiler(str(file1_path))
                label = "Profiler"
        except Exception as e:
            print(f"Error extracting layers: {e}", file=sys.stderr)
            return 1

        print(f"{label} CSV: {file1_path}")
        print(f"Loaded {len(layers)} {label.lower()} layers")

        if args.filter_optype:
            filter_norm = normalize_optype(args.filter_optype)
            layers = [layer for layer in layers if normalize_optype(layer['optype']) == filter_norm]
            print(f"Filtered to {len(layers)} layers with optype='{args.filter_optype}'")

        if args.summarize_by_signature:
            _print_signature_summary(
                layers,
                label,
                args.strip_leading_ones,
                args.strip_singleton_dims,
                include_perf=args.perf,
            )
        if args.perf:
            if args.summarize_by_signature:
                _print_perf_standalone_by_signature(
                    layers,
                    label,
                    args.strip_leading_ones,
                    args.strip_singleton_dims,
                )
            else:
                _print_perf_standalone(layers, label)

        if args.xlsx:
            try:
                _write_xlsx_report(
                    args.xlsx,
                    polaris_layers=layers if ftype == 'polaris' else None,
                    profiler_layers=layers if ftype == 'profiler' else None,
                    stats=None,
                    polaris_label="Polaris",
                    profiler_label="Profiler",
                    polaris_source=args.file1 if ftype == 'polaris' else None,
                    profiler_source=args.file1 if ftype == 'profiler' else None,
                    strip_leading_ones=args.strip_leading_ones,
                    strip_singleton_dims=args.strip_singleton_dims,
                )
                print(f"\nWrote XLSX report: {args.xlsx}")
            except Exception as e:
                print(f"Error writing XLSX report: {e}", file=sys.stderr)
                return 1
        return 0

    # --- Two-file mode ---
    type1 = detect_file_type(str(file1_path))
    type2 = detect_file_type(str(file2_path))

    if type1 is None or type2 is None:
        print("Error: Could not determine file types. Expected CSV files with "
              "'archname' (polaris) or 'OP CODE' (profiler) columns.", file=sys.stderr)
        return 1

    if type1 == type2:
        print(f"Error: Both files appear to be {type1} CSVs. "
              f"Expected one polaris and one profiler CSV.", file=sys.stderr)
        return 1

    # Assign files based on type
    polaris_file = str(file1_path) if type1 == 'polaris' else str(file2_path)
    profiler_file = str(file1_path) if type1 == 'profiler' else str(file2_path)

    print(f"Polaris CSV: {polaris_file}")
    print(f"Profiler CSV: {profiler_file}")
    print()

    # Extract layers
    try:
        polaris_layers = layers_polaris(polaris_file)
        profiler_layers = layers_profiler(profiler_file)
    except Exception as e:
        print(f"Error extracting layers: {e}", file=sys.stderr)
        return 1

    print(f"Loaded {len(polaris_layers)} polaris layers, {len(profiler_layers)} profiler layers")

    # Filter by optype if requested
    if args.filter_optype:
        filter_optype_norm = normalize_optype(args.filter_optype)

        polaris_layers = [
            layer for layer in polaris_layers
            if normalize_optype(layer['optype']) == filter_optype_norm
        ]
        profiler_layers = [
            layer for layer in profiler_layers
            if normalize_optype(layer['optype']) == filter_optype_norm
        ]

        print(f"Filtered to {len(polaris_layers)} polaris layers, {len(profiler_layers)} profiler layers with optype='{args.filter_optype}'")

        if len(polaris_layers) == 0 and len(profiler_layers) == 0:
            print(f"Warning: No layers found with optype='{args.filter_optype}'")
            return 0

    print()

    # Compare layers (shape / attribute matching)
    stats = compare_layers(polaris_layers, profiler_layers, args.max_search_distance,
                           args.strip_leading_ones, args.strip_singleton_dims,
                           ignore_attrs=args.ignore_attrs)

    # Print shape-comparison summary
    print_summary(stats)

    if args.summarize_by_signature:
        _print_signature_summary(
            polaris_layers,
            "Polaris",
            args.strip_leading_ones,
            args.strip_singleton_dims,
            include_perf=args.perf,
        )
        _print_signature_summary(
            profiler_layers,
            "Profiler",
            args.strip_leading_ones,
            args.strip_singleton_dims,
            include_perf=args.perf,
        )

    # Performance comparison (when --perf is enabled)
    if args.perf:
        _print_perf_comparison(
            profiler_layers,
            polaris_layers,
            by_signature=args.summarize_by_signature,
            strip_leading_ones=args.strip_leading_ones,
            strip_singleton_dims=args.strip_singleton_dims,
        )

    if args.xlsx:
        try:
            _write_xlsx_report(
                args.xlsx,
                polaris_layers=polaris_layers,
                profiler_layers=profiler_layers,
                stats=stats,
                polaris_label="Polaris",
                profiler_label="Profiler",
                polaris_source=polaris_file,
                profiler_source=profiler_file,
                strip_leading_ones=args.strip_leading_ones,
                strip_singleton_dims=args.strip_singleton_dims,
            )
            print(f"\nWrote XLSX report: {args.xlsx}")
        except Exception as e:
            print(f"Error writing XLSX report: {e}", file=sys.stderr)
            return 1

    return 0


if __name__ == '__main__':
    sys.exit(main())
