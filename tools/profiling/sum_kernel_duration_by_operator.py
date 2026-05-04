#!/usr/bin/env python3
# SPDX-FileCopyrightText: (C) 2026 Tenstorrent AI ULC
# SPDX-License-Identifier: Apache-2.0

"""
Sum kernel device durations by operator type from profiler CSV/XLSX tables.

Rules for operator type:
- Default: value in "OP CODE" column
- If "OP CODE" == "BinaryNgDeviceOperation": derive from "binary_op_type" in "ATTRIBUTES"
- If "OP CODE" == "UnaryDeviceOperation": derive from "unary_op_type" in "ATTRIBUTES"

Duration conversion:
- Input column "device kernel duration [ns]" is interpreted as nanoseconds
- Output totals are in milliseconds
"""

from __future__ import annotations

import argparse
import ast
import csv
import math
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import yaml


def _parse_attributes(raw: Any) -> dict[str, Any]:
    if raw is None:
        return {}
    text = str(raw).strip()
    if not text or text.lower() in {"nan", "none", "nat"}:
        return {}

    try:
        parsed = yaml.safe_load(text.replace(";", ","))
        if isinstance(parsed, dict):
            return parsed
    except Exception:
        pass

    try:
        parsed = ast.literal_eval(text)
        if isinstance(parsed, dict):
            return parsed
    except (ValueError, SyntaxError, TypeError):
        pass

    return {}


def _parse_duration_ns(raw: Any) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, float) and math.isnan(raw):
        return None
    text = str(raw).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _derive_operator_type(op_code: str, attributes_raw: Any) -> str:
    attrs = _parse_attributes(attributes_raw)

    if op_code == "BinaryNgDeviceOperation":
        return str(attrs.get("binary_op_type") or op_code)

    if op_code == "UnaryDeviceOperation":
        return str(attrs.get("unary_op_type") or op_code)

    return op_code


def _normalize_header(name: str) -> str:
    # Match headers robustly across casing and arbitrary whitespace.
    return re.sub(r"\s+", " ", name.strip()).lower()


def _canonicalize_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    required_columns = ["OP CODE", "ATTRIBUTES", "device kernel duration [ns]"]
    required_norm = {_normalize_header(col): col for col in required_columns}

    header_map: dict[str, str] = {}
    for header in rows[0].keys():
        if header is None:
            continue
        normalized = _normalize_header(header)
        if normalized in required_norm:
            header_map[required_norm[normalized]] = str(header)

    missing = [col for col in required_columns if col not in header_map]
    if missing:
        raise KeyError(
            "Missing required columns (case-insensitive, whitespace-tolerant): "
            + ", ".join(missing)
        )

    canonical_rows: list[dict[str, Any]] = []
    for row in rows:
        canonical_rows.append({col: row.get(source_col) for col, source_col in header_map.items()})
    return canonical_rows


def _load_rows(input_path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        with input_path.open(newline="", encoding="utf-8") as handle:
            return list(csv.DictReader(handle))

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        try:
            import openpyxl
        except Exception as exc:
            raise RuntimeError(
                "Reading Excel files requires openpyxl. Install with: pip install openpyxl"
            ) from exc

        wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                available = ", ".join(wb.sheetnames)
                raise KeyError(f"Sheet '{sheet_name}' not found. Available sheets: {available}")
            ws = wb[sheet_name]
        else:
            _active = wb.active
            if _active is None:
                return []
            ws = _active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return []
        header_names = [str(h).strip() if h is not None else "" for h in headers]
        rows: list[dict[str, Any]] = []
        for values in rows_iter:
            row = {header_names[i]: values[i] for i in range(min(len(header_names), len(values)))}
            rows.append(row)
        return rows

    raise ValueError(f"Unsupported input format: {input_path.suffix}")


def summarize_kernel_duration_ms(rows: list[dict[str, Any]]) -> dict[str, tuple[int, float]]:
    canonical_rows = _canonicalize_rows(rows)
    if not canonical_rows:
        return {}

    totals_ms: dict[str, float] = defaultdict(float)
    counts: dict[str, int] = defaultdict(int)
    for row in canonical_rows:
        op_code = str(row.get("OP CODE", "")).strip()
        if not op_code:
            continue
        operator_type = _derive_operator_type(op_code, row.get("ATTRIBUTES"))

        duration_ns = _parse_duration_ns(row.get("device kernel duration [ns]"))
        if duration_ns is None:
            continue
        totals_ms[operator_type] += duration_ns / 1_000_000.0
        counts[operator_type] += 1
    return {op: (counts[op], total_ms) for op, total_ms in totals_ms.items()}


def _write_csv(output_path: Path, totals_ms: dict[str, tuple[int, float]]) -> None:
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["operator_type", "entry_count", "kernel_device_duration_ms"])
        for operator_type, (count, total_ms) in sorted(
            totals_ms.items(), key=lambda kv: kv[1][1], reverse=True
        ):
            writer.writerow([operator_type, count, f"{total_ms:.6f}"])


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Sum kernel device duration [ns] to ms grouped by operator type."
    )
    parser.add_argument("input_file", type=Path, help="Profiler table file (.csv or .xlsx)")
    parser.add_argument(
        "--sheet-name",
        type=str,
        default=None,
        help="Excel sheet name to read. Defaults to workbook active sheet.",
    )
    parser.add_argument(
        "-o",
        "--output-csv",
        type=Path,
        default=None,
        help="Optional output CSV path. If omitted, prints to stdout.",
    )
    args = parser.parse_args()

    rows = _load_rows(args.input_file, sheet_name=args.sheet_name)
    totals_ms = summarize_kernel_duration_ms(rows)
    grand_total_ms = sum(total_ms for _, total_ms in totals_ms.values())
    grand_total_count = sum(count for count, _ in totals_ms.values())

    if args.output_csv is not None:
        _write_csv(args.output_csv, totals_ms)
        print(f"Wrote {len(totals_ms)} rows to {args.output_csv}")
        print(f"total_entries,{grand_total_count}")
        print(f"total_kernel_device_duration_ms,{grand_total_ms:.6f}")
        return 0

    print("operator_type,entry_count,kernel_device_duration_ms")
    for operator_type, (count, total_ms) in sorted(totals_ms.items(), key=lambda kv: kv[1][1], reverse=True):
        print(f"{operator_type},{count},{total_ms:.6f}")
    print(f"total_entries,{grand_total_count}")
    print(f"total_kernel_device_duration_ms,{grand_total_ms:.6f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
