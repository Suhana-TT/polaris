# SPDX-FileCopyrightText: (C) 2025 Tenstorrent AI ULC
# SPDX-License-Identifier: Apache-2.0
"""
Chart Builder Module - Create visualization charts for correlation analysis.

This module handles creation of charts and visualizations for correlation
results, including the S-curve analysis chart.

Responsibilities:
-----------------
- Create S-curve analysis charts
- Configure chart axes and ticks
- Calculate chart statistics (min, max, median, geomean)
- Format chart worksheets
- Add charts to Excel workbooks

Classes:
--------
- ScurveChartBuilder: Builder for S-curve analysis charts
  - prepare_data(): Extract and sort ratio data
  - calculate_statistics(): Calculate min, max, median, geomean
  - create_worksheet(): Create worksheet with title and stats
  - add_data_table(): Add sorted data table
  - create_chart(): Create and configure line chart
  - build(): Build complete S-curve sheet in workbook

Functions:
----------
- add_scurve_chart() - High-level facade for chart creation

Migration Status:
-----------------
🟢 PARTIALLY IMPLEMENTED - Phase 3 in progress
  ✅ ScurveChartBuilder class created
  ✅ add_scurve_chart() function implemented
  🔴 add_scurve_sheet() still in run_ttsi_corr.py (will update to use builder)

Usage:
------
    from ttsi_corr.chart_builder import add_scurve_chart
    
    add_scurve_chart(
        workbook=wb,
        comparison=correlation_data
    )

See Also:
---------
- tools/run_ttsi_corr.py: Main correlation script
- tools/ttsi_corr/excel_writer.py: Table generation
- openpyxl.chart: Chart library used for implementation
"""

import math
from typing import Any, Optional

from loguru import logger
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet


class ScurveChartBuilder:
    """
    Builder for S-curve analysis charts.
    
    This class implements the Builder pattern for creating complex S-curve
    charts. It separates the construction of a complex chart into distinct
    steps, making it easier to test, maintain, and extend.
    
    Attributes:
        comparison (list): Original correlation data
        ratio_data (list): Extracted and sorted ratio data
        statistics (dict): Calculated statistics (min, max, median, geomean)
        ws: Worksheet object (created during build)
        chart: LineChart object (created during build)
    
    Example:
        >>> builder = ScurveChartBuilder(comparison_data)
        >>> builder.build(workbook)
        # Creates 'S-Curve Analysis' sheet with chart
    """
    
    def __init__(self, comparison: list[dict[str, Any]]):
        """
        Initialize ScurveChartBuilder with correlation data.
        
        Args:
            comparison: List of correlation data dictionaries.
                       Each dict must have 'Ratio-HLM-to-SiTarget' key.
        """
        self.comparison = comparison
        self.ratio_data: list[dict[str, Any]] = []
        self.statistics: dict[str, float] = {}
        self.ws: Optional[Worksheet] = None
        self.chart: Optional[LineChart] = None
    
    def prepare_data(self) -> 'ScurveChartBuilder':
        """
        Extract and sort ratio data from comparison.
        
        Extracts 'Ratio-HLM-to-SiTarget' values along with workload info,
        filters out invalid values, and sorts by ratio.
        
        Returns:
            Self for method chaining
            
        Side Effects:
            Populates self.ratio_data with sorted ratio information
        """
        self.ratio_data = []
        for row in self.comparison:
            ratio = row.get('Ratio-HLM-to-SiTarget')
            if ratio is not None and isinstance(ratio, (int, float)):
                self.ratio_data.append({
                    'Workload': row.get('Workload', 'Unknown'),
                    'Batch': row.get('Batch', ''),
                    'Ratio': float(ratio)
                })
        
        if not self.ratio_data:
            logger.warning('No valid ratio data found for S-curve analysis')
            return self
        
        # Sort by ratio value (ascending)
        self.ratio_data.sort(key=lambda x: x['Ratio'])
        logger.debug('Prepared {} ratio data points', len(self.ratio_data))
        
        return self
    
    def calculate_statistics(self) -> 'ScurveChartBuilder':
        """
        Calculate summary statistics for ratio data.
        
        Computes min, max, median, and geometric mean of ratios.
        
        Returns:
            Self for method chaining
            
        Side Effects:
            Populates self.statistics with calculated values
            
        Raises:
            ValueError: If ratio_data is empty
        """
        if not self.ratio_data:
            raise ValueError('Cannot calculate statistics: ratio_data is empty')
        
        ratios = [d['Ratio'] for d in self.ratio_data]
        self.statistics = {
            'min': min(ratios),
            'max': max(ratios),
            'median': ratios[len(ratios) // 2],
            'geomean': math.exp(sum(math.log(x) for x in ratios) / len(ratios))
        }
        
        logger.debug('Statistics: min={:.4f}, max={:.4f}, median={:.4f}, geomean={:.4f}',
                    self.statistics['min'], self.statistics['max'],
                    self.statistics['median'], self.statistics['geomean'])
        
        return self
    
    def create_worksheet(self, wb: Workbook) -> 'ScurveChartBuilder':
        """
        Create worksheet with title and statistics summary.
        
        Args:
            wb: Workbook object to add the sheet to
            
        Returns:
            Self for method chaining
            
        Side Effects:
            Creates new worksheet named 'S-Curve Analysis'
            Adds title and statistics table
            Sets self.ws to the created worksheet
        """
        ws = wb.create_sheet(title='S-Curve Analysis')
        self.ws = ws
        
        # Add title
        ws['A1'] = 'S-Curve Analysis: HLM vs Silicon Target Correlation'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Add statistics summary
        ws['A3'] = 'Statistics Summary'
        ws['A3'].font = Font(bold=True)
        ws['A4'] = 'Total Workloads:'
        ws['B4'] = len(self.ratio_data)
        ws['A5'] = 'Minimum Ratio:'
        ws['B5'] = self.statistics['min']
        ws['B5'].number_format = '0.0000'
        ws['A6'] = 'Maximum Ratio:'
        ws['B6'] = self.statistics['max']
        ws['B6'].number_format = '0.0000'
        ws['A7'] = 'Median Ratio:'
        ws['B7'] = self.statistics['median']
        ws['B7'].number_format = '0.0000'
        ws['A8'] = 'Geometric Mean:'
        ws['B8'] = self.statistics['geomean']
        ws['B8'].number_format = '0.0000'
        ws['B8'].font = Font(bold=True)
        
        logger.debug('Created worksheet with statistics')
        
        return self
    
    def add_data_table(self) -> 'ScurveChartBuilder':
        """
        Add data table with sorted ratio data.
        
        Returns:
            Self for method chaining
            
        Side Effects:
            Adds data table headers and rows to worksheet
            
        Raises:
            ValueError: If worksheet not created (call create_worksheet first)
        """
        if self.ws is None:
            raise ValueError('Worksheet not created. Call create_worksheet() first.')
        
        ws = self.ws  # Local variable for type narrowing
        
        # Add data table headers
        ws['A10'] = 'Index'
        ws['B10'] = 'Workload'
        ws['C10'] = 'Batch'
        ws['D10'] = 'Ratio-HLM-to-SiTarget'
        for cell in ['A10', 'B10', 'C10', 'D10']:
            ws[cell].font = Font(bold=True)
            ws[cell].alignment = Alignment(horizontal='center')
        
        # Add data rows
        for idx, data in enumerate(self.ratio_data, start=1):
            row_num = 10 + idx
            ws[f'A{row_num}'] = idx
            ws[f'B{row_num}'] = data['Workload']
            ws[f'C{row_num}'] = data['Batch']
            ws[f'D{row_num}'] = data['Ratio']
            ws[f'D{row_num}'].number_format = '0.0000'
        
        logger.debug('Added data table with {} rows', len(self.ratio_data))
        
        return self
    
    def create_chart(self) -> 'ScurveChartBuilder':
        """
        Create and configure S-curve line chart.
        
        Creates a line chart with:
        - Title and styling
        - Configured X and Y axes with ticks
        - Smooth line for S-curve appearance
        - Data series from worksheet
        
        Returns:
            Self for method chaining
            
        Side Effects:
            Creates LineChart object and configures it
            Sets self.chart to the created chart
            
        Raises:
            ValueError: If worksheet not created
        """
        if self.ws is None:
            raise ValueError('Worksheet not created. Call create_worksheet() first.')
        
        ws = self.ws  # Local variable for type narrowing
        
        # Create line chart
        chart = LineChart()
        self.chart = chart
        chart.title = 'S-Curve: HLM-to-Silicon Target Ratio Distribution'
        chart.style = 12  # Modern style
        chart.height = 12  # Larger chart
        chart.width = 20
        
        # Configure X-axis (Workload Index)
        chart.x_axis.delete = False
        chart.x_axis.title = 'Workload Index (sorted by ratio)'
        chart.x_axis.tickLblPos = 'low'
        
        # Calculate appropriate X-axis tick interval (~10 ticks)
        x_tick_interval = max(1, len(self.ratio_data) // 10)
        chart.x_axis.majorUnit = x_tick_interval  # type: ignore[attr-defined]
        chart.x_axis.minorUnit = x_tick_interval  # type: ignore[attr-defined]
        
        # Configure Y-axis (Ratio values)
        chart.y_axis.delete = False
        chart.y_axis.title = 'Ratio-HLM-to-SiTarget'
        chart.y_axis.tickLblPos = 'high'
        
        # Set Y-axis range with padding
        y_min = max(0.0, self.statistics['min'] - 0.1)
        y_max = self.statistics['max'] + 0.1
        chart.y_axis.scaling.min = y_min
        chart.y_axis.scaling.max = y_max
        
        # Calculate Y-axis tick interval (~10 ticks)
        y_range = y_max - y_min
        y_tick_interval = round(y_range / 10, 2)
        if y_tick_interval > 0:
            chart.y_axis.majorUnit = y_tick_interval
        
        # Add data series
        data_ref = Reference(ws, min_col=4, min_row=10, max_row=10 + len(self.ratio_data))
        cats_ref = Reference(ws, min_col=1, min_row=11, max_row=10 + len(self.ratio_data))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        # Configure series (smooth line)
        if chart.series:
            series = chart.series[0]
            series.smooth = True  # Smooth line for S-curve appearance
        
        logger.debug('Created chart with data series')
        
        return self
    
    def build(self, wb: Workbook) -> None:
        """
        Build complete S-curve analysis sheet with chart.
        
        Executes all steps in order:
        1. Prepare data (extract and sort ratios)
        2. Calculate statistics
        3. Create worksheet
        4. Add data table
        5. Create chart
        6. Add chart to sheet
        7. Adjust column widths
        
        Args:
            wb: Workbook object to add the sheet to
            
        Side Effects:
            Creates complete 'S-Curve Analysis' worksheet in workbook
            
        Note:
            This is the main entry point. Call this method to create
            the entire S-curve sheet in one go.
        """
        logger.debug('Building S-curve analysis sheet')
        
        # Execute build steps
        (self.prepare_data()
             .calculate_statistics()
             .create_worksheet(wb)
             .add_data_table()
             .create_chart())
        
        # Add chart to sheet
        if self.ws and self.chart:
            self.ws.add_chart(self.chart, 'F3')
        
        # Adjust column widths
        if self.ws:
            self.ws.column_dimensions['A'].width = 8
            self.ws.column_dimensions['B'].width = 30
            self.ws.column_dimensions['C'].width = 10
            self.ws.column_dimensions['D'].width = 20
        
        logger.info('S-curve analysis sheet created with {} data points', len(self.ratio_data))


def add_scurve_chart(wb: Workbook, comparison: list[dict[str, Any]]) -> None:
    """
    Add S-curve analysis sheet with chart to workbook (HIGH-LEVEL FACADE).
    
    Creates a complete S-curve analysis sheet including:
    - Summary statistics (min, max, median, geometric mean)
    - Sorted data table
    - Embedded line chart showing S-curve distribution
    
    Args:
        wb: Workbook object to add the chart sheet to
        comparison: List of dictionaries containing correlation data.
                   Each dict must have 'Ratio-HLM-to-SiTarget' key.
        
    Side Effects:
        Adds a new worksheet named 'S-Curve Analysis' to the workbook
        
    Example:
        >>> from openpyxl import Workbook
        >>> from ttsi_corr.chart_builder import add_scurve_chart
        >>> 
        >>> wb = Workbook()
        >>> comparison = [
        ...     {'Workload': 'BERT', 'Batch': 16, 'Ratio-HLM-to-SiTarget': 0.95},
        ...     {'Workload': 'ResNet', 'Batch': 32, 'Ratio-HLM-to-SiTarget': 1.05}
        ... ]
        >>> add_scurve_chart(wb, comparison)
        >>> wb.save('output.xlsx')
        
    Note:
        Migrated from tools.run_ttsi_corr.add_scurve_sheet() in Phase 3.
        Uses ScurveChartBuilder class internally.
    """
    builder = ScurveChartBuilder(comparison)
    builder.build(wb)

