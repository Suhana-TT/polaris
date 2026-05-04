#!/usr/bin/env python3
# SPDX-FileCopyrightText: (C) 2025 Tenstorrent AI ULC
# SPDX-License-Identifier: Apache-2.0

"""
Build a master dictionary from Excel or CSV: key tuple -> flat single (``num_cores`` + stats) or curve equations (curve).
After load, OP CODE cells are normalized to Polaris-style layer *types* (e.g. matmul, add, mul, tilize);
``TilizeWithValPadding*`` → ``tilizewithvalpadding`` (other ``Tilize*`` → ``tilize``); ``UntilizeWithUnpadding*``
→ ``untilizewithunpadding`` (other ``Untilize*`` → ``untilize``). ``BinaryNgDeviceOperation`` and ``UnaryDeviceOperation`` rows use
``ATTRIBUTES`` (``binary_op_type`` / ``UnaryOpType::…`` in ``op_chain``) when present; all keys, grouping, and mode checks use these types — not
full profiler class names. Key tuple length is **9** if all INPUT_1_* and INPUT_2_* are blank, **16** if
INPUT_1_* is all set and INPUT_2_* all blank, **23** if INPUT_1_* and INPUT_2_* are all set (ternary ops
e.g. LayerNorm). The key includes ``MATH FIDELITY`` after the INPUT_0 slot. Matmul uses ``entry_type:
hybrid`` in YAML (``single`` + ``curve`` branches) when combining model and sweep data; see
``doc/YAML_MASTER_FORMAT.md``. CLI: repeatable ``--model-run`` and ``--sweep-run`` to merge Excel or CSV in one
invocation (CSV is read with stdlib ``csv``; Excel ``.xlsx`` / ``.xlsm`` with **openpyxl**). Writes ``schema_name`` ``correqn.tt-perf-master``, ``schema_version`` (``tt_perf_master_schema.MASTER_YAML_SCHEMA_VERSION``, **1** until first release).

Run from repository root: ``python -m tools.perf_lookup.tt_perf_mapper …``, or ``python tools/perf_lookup/tt_perf_mapper.py …`` (bootstrap adds repo root to ``sys.path``).
"""
from __future__ import annotations

import sys
from pathlib import Path

_mp = Path(__file__).resolve()
if (
    len(_mp.parents) >= 3
    and _mp.parent.name == "perf_lookup"
    and _mp.parents[1].name == "tools"
):
    _repo_root = str(_mp.parents[2])
    if _repo_root not in sys.path:
        sys.path.insert(0, _repo_root)

import argparse
import csv
import math
import re
from collections import defaultdict
from collections.abc import Mapping
from dataclasses import dataclass
from typing import Any
from zipfile import BadZipFile

import numpy as np
import yaml
from loguru import logger
from scipy.optimize import curve_fit  # type: ignore[import-untyped]

from tools.profiling.op_canonical import normalize_profiler_opcode

from tools.perf_lookup.tt_perf_master_loader import load_existing_yaml
from tools.perf_lookup.tt_perf_master_schema import (
    MASTER_CURVE_FAMILY_KEY,
    MASTER_CURVE_FAMILY_LINEAR,
    MASTER_CURVE_FAMILY_POWER,
    MASTER_DURATION_MS_KEY,
    MASTER_ENTRY_TYPE_CURVE,
    MASTER_ENTRY_TYPE_HYBRID,
    MASTER_ENTRY_TYPE_KEY,
    MASTER_ENTRY_TYPE_SINGLE,
    MASTER_HYBRID_CURVE_KEY,
    MASTER_HYBRID_SINGLE_KEY,
    MASTER_SINGLE_NUM_CORES_KEY,
    MASTER_SINGLE_STAT_KEYS,
    MASTER_SINGLE_STAT_KEYS_SET,
    normalize_flat_single_payload,
    MASTER_YAML_ENTRIES_KEY,
    MASTER_YAML_ENTRY_VALUE_FIELD,
    MASTER_YAML_RECORD_KEY_FIELD,
    MASTER_YAML_SCHEMA_NAME,
    MASTER_YAML_SCHEMA_NAME_KEY,
    MASTER_YAML_SCHEMA_VERSION,
    MASTER_YAML_SCHEMA_VERSION_KEY,
    tuple_to_labeled_key_map,
    MATH_FIDELITY_NA,
)


def is_na(v: Any) -> bool:
    """True for ``None`` and floating NaN (spreadsheet empty cells are usually ``None``)."""
    if v is None:
        return True
    if isinstance(v, (float, np.floating)):
        try:
            return math.isnan(float(v))
        except (TypeError, ValueError):
            return False
    return False


@dataclass
class InputTable:
    """Column names in file order; each row maps column name to cell value."""

    columns: list[str]
    rows: list[dict[str, Any]]

    @property
    def column_set(self) -> set[str]:
        return set(self.columns)


# Key tuple columns in order: OP CODE, INPUT_0 (7), MATH FIDELITY, INPUT_1 (7), INPUT_2 (7).
# See ``build_key_tuple``.
KEY_TUPLE_PREFIX_LEN = 1 + 7 + 1  # OP CODE + INPUT_0 + MATH FIDELITY
KEY_TUPLE_INPUT_SLOT_LEN = 7  # fields per logical input (pads + layout + datatype + memory)
_MATH_FIDELITY_KEY_IDX = 8  # position of MATH FIDELITY in the key tuple

# Op codes for which the TTNN API caller controls math fidelity via
# ``compute_kernel_config``.  For all other ops the hardware kernel
# determines fidelity internally; the LUT key stores ``N/A`` so that
# the runtime default (also ``N/A``) matches unconditionally.
_MATH_FIDELITY_CALLER_CONTROLLED_OPS = frozenset({"layernorm"})
KEY_TUPLE_COLUMN_NAMES = [
    "OP CODE",
    "INPUT_0_W_PAD[LOGICAL]",
    "INPUT_0_Z_PAD[LOGICAL]",
    "INPUT_0_Y_PAD[LOGICAL]",
    "INPUT_0_X_PAD[LOGICAL]",
    "INPUT_0_LAYOUT",
    "INPUT_0_DATATYPE",
    "INPUT_0_MEMORY",
    "MATH FIDELITY",
    "INPUT_1_W_PAD[LOGICAL]",
    "INPUT_1_Z_PAD[LOGICAL]",
    "INPUT_1_Y_PAD[LOGICAL]",
    "INPUT_1_X_PAD[LOGICAL]",
    "INPUT_1_LAYOUT",
    "INPUT_1_DATATYPE",
    "INPUT_1_MEMORY",
    "INPUT_2_W_PAD[LOGICAL]",
    "INPUT_2_Z_PAD[LOGICAL]",
    "INPUT_2_Y_PAD[LOGICAL]",
    "INPUT_2_X_PAD[LOGICAL]",
    "INPUT_2_LAYOUT",
    "INPUT_2_DATATYPE",
    "INPUT_2_MEMORY",
]

# Excel column (nanoseconds). Master YAML duration key is ``MASTER_DURATION_MS_KEY`` (imported from schema).
EXCEL_DURATION_COLUMN = "DEVICE KERNEL DURATION [ns]"
NS_PER_MS = 1_000_000.0

# Excel stat columns (input). Master YAML uses ``OUTPUT_*`` keys below.
EXCEL_COL_DRAM_BW_UTIL = "DRAM BW UTIL (%)"
EXCEL_COL_NOC_UTIL = "NOC UTIL (%)"
EXCEL_COL_MULTICAST_NOC = "MULTICAST NOC UTIL (%)"
EXCEL_COL_NPE_CONG = "NPE CONG IMPACT (%)"
EXCEL_COL_SFPU = "SFPU Util Median (%)"
EXCEL_COL_FPU = "FPU Util Median (%)"

OUTPUT_KEY_MEM_UTIL = "mem_util"
OUTPUT_KEY_NOC_UTIL = "noc_util"
OUTPUT_KEY_NOC_MULTICAST_UTIL = "noc_multicast_util"
OUTPUT_KEY_NPE_CONG_IMPACT_PCT = "npe_cong_impact_pct"
OUTPUT_KEY_VECTOR_PIPE_UTIL = "vector_pipe_util"
OUTPUT_KEY_MATRIX_PIPE_UTIL = "matrix_pipe_util"

STATS_COLUMNS_REQUIRED = [
    EXCEL_DURATION_COLUMN,
    EXCEL_COL_DRAM_BW_UTIL,
    EXCEL_COL_NOC_UTIL,
    EXCEL_COL_NPE_CONG,
    EXCEL_COL_SFPU,
    EXCEL_COL_FPU,
]
STATS_COLUMN_OPTIONAL = EXCEL_COL_MULTICAST_NOC

# Merged three-way ops CSV (e.g. ``ops_perf_three_csv_merge`` output): util columns are suffixed.
_MERGED_OPS_STAT_ALIASES: tuple[tuple[str, str], ...] = (
    (EXCEL_COL_DRAM_BW_UTIL, f"{EXCEL_COL_DRAM_BW_UTIL}_noctrace"),
    (EXCEL_COL_NOC_UTIL, f"{EXCEL_COL_NOC_UTIL}_noctrace"),
    (EXCEL_COL_NPE_CONG, f"{EXCEL_COL_NPE_CONG}_noctrace"),
    (EXCEL_COL_SFPU, f"{EXCEL_COL_SFPU}_fpuutil"),
    (EXCEL_COL_FPU, f"{EXCEL_COL_FPU}_fpuutil"),
    (STATS_COLUMN_OPTIONAL, f"{STATS_COLUMN_OPTIONAL}_noctrace"),
)

CV_THRESHOLD = 0.10

# Iglewicz–Hoaglin modified Z (MAD-based); drop |M*| > threshold before CV. n < 3 or MAD==0 → no rejection.
CV_OUTLIER_MODIFIED_Z_THRESHOLD = 3.5

# First key-tuple field for matmul after Polaris layer-type normalization.
POLARIS_LAYER_MATMUL = "matmul"


def _single_mode_core_count_key(core_count: float, key_t: tuple) -> int:
    """Whole CORE COUNT → int master dict key; non-integral values are rounded (warning)."""
    x = float(core_count)
    if not x.is_integer():
        logger.warning(
            "CORE COUNT {} is not a whole number; using int(round(...)) for single-mode "
            "master key; key_tuple={}",
            core_count,
            list(key_t),
        )
        return int(round(x))
    return int(x)


EXCEL_ATTRIBUTES_COLUMN = "ATTRIBUTES"

# PAD column suffix pattern: value like "224[224]" -> use 224
PAD_PATTERN = re.compile(r"^\s*(\d+)\s*\[\s*(\d+)\s*\]\s*$")


def _parse_attributes_cell(raw) -> dict | None:
    """Parse an ``ATTRIBUTES`` cell into a dict, or ``None``."""
    if raw is None or is_na(raw):
        return None
    s = str(raw).strip()
    if not s:
        return None
    try:
        parsed = yaml.safe_load(s.replace(";", ","))
    except yaml.YAMLError:
        return None
    return parsed if isinstance(parsed, dict) else None


def apply_polaris_layer_type_column(table: InputTable) -> None:
    """In-place: replace ``OP CODE`` with Polaris layer *types* for downstream logic and YAML keys.

    Delegates to :func:`tools.profiling.op_canonical.normalize_profiler_opcode`.
    """
    if "OP CODE" not in table.column_set:
        return
    has_attrs = EXCEL_ATTRIBUTES_COLUMN in table.column_set
    for row in table.rows:
        attrs = _parse_attributes_cell(row.get(EXCEL_ATTRIBUTES_COLUMN)) if has_attrs else None
        row["OP CODE"] = normalize_profiler_opcode(row.get("OP CODE", ""), attrs)


def parse_input_spec(input_str: str) -> tuple[str | None, str]:
    """Return (sheet_name or None, file_path)."""
    s = input_str.strip()
    if "@" in s:
        sheet, path = s.split("@", 1)
        return sheet.strip() or None, path.strip()
    return None, s.strip()


def load_excel(sheet_name: str | None, excel_path: Path) -> InputTable:
    """Load one sheet: either sheet_name or the only sheet. Multi-sheet without sheet_name -> error."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        if sheet_name is not None:
            if sheet_name not in names:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. Available: {names}"
                )
            sn = sheet_name
        else:
            if len(names) != 1:
                raise ValueError(
                    f"File has {len(names)} sheets; pass sheet@file to select one. Sheets: {names}"
                )
            sn = names[0]
        ws = wb[sn]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return InputTable([], [])
        raw_names = [str(c).strip() if c is not None else "" for c in header]
        col_pairs = [(i, n) for i, n in enumerate(raw_names) if n]
        columns = [n for _, n in col_pairs]
        data_rows: list[dict[str, Any]] = []
        for raw_row in rows_iter:
            d: dict[str, Any] = {}
            for i, name in col_pairs:
                val = raw_row[i] if i < len(raw_row) else None
                d[name] = val
            data_rows.append(d)
        return InputTable(columns, data_rows)
    finally:
        wb.close()


def load_csv(csv_path: Path, *, encoding: str = "utf-8") -> InputTable:
    """Load a CSV into an :class:`InputTable` using stdlib ``csv.reader`` (newline=\"\", RFC-style parsing)."""
    with open(csv_path, newline="", encoding=encoding) as f:
        reader = csv.reader(f)
        try:
            header_row = next(reader)
        except StopIteration:
            return InputTable([], [])
        names = [c.strip() for c in header_row]
        if not names:
            return InputTable([], [])
        n = len(names)
        col_pairs = [(i, name) for i, name in enumerate(names) if name]
        columns = [name for _, name in col_pairs]
        data_rows: list[dict[str, Any]] = []
        for row in reader:
            if len(row) < n:
                row = row + [""] * (n - len(row))
            elif len(row) > n:
                row = row[:n]
            d = {name: row[i] if i < len(row) else "" for i, name in col_pairs}
            data_rows.append(d)
    return InputTable(columns, data_rows)


def load_input_table(sheet_name: str | None, path: Path) -> InputTable:
    """Load ``.csv`` (stdlib csv) or Excel ``.xlsx`` / ``.xlsm``. Sheet selection applies only to Excel."""
    suf = path.suffix.lower()
    if suf == ".csv":
        if sheet_name is not None:
            raise ValueError(
                f"Sheet selection is only valid for Excel files, not CSV: {path}"
            )
        return load_csv(path)
    if suf in (".xlsx", ".xlsm"):
        return load_excel(sheet_name, path)
    raise ValueError(
        f"Unsupported input file type {path.suffix!r}; use .csv, .xlsx, or .xlsm: {path}"
    )


def apply_merged_ops_stat_aliases(table: InputTable) -> None:
    """If canonical stat columns are missing but merged ``*_noctrace`` / ``*_fpuutil`` exist, copy into canonical names."""
    cs = table.column_set
    for canon, alt in _MERGED_OPS_STAT_ALIASES:
        if canon in cs:
            continue
        if alt in cs:
            table.columns.append(canon)
            cs.add(canon)
            for row in table.rows:
                row[canon] = row.get(alt)


def resolve_core_col(table: InputTable) -> str:
    if "CORE COUNT" in table.column_set:
        return "CORE COUNT"
    if "Core Count" in table.column_set:
        return "Core Count"
    raise KeyError("CORE COUNT column not found. Available: " + str(list(table.columns)))


def validate_columns(
    table: InputTable,
) -> tuple[str, str, list[str], list[str], list[str]]:
    """Return (core_col, duration_col, key_cols, stat_cols_required, stat_cols_optional)."""
    core_col = resolve_core_col(table)
    duration_col = EXCEL_DURATION_COLUMN
    cs = table.column_set
    missing_key = [c for c in KEY_TUPLE_COLUMN_NAMES if c not in cs]
    if missing_key:
        raise KeyError(
            f"Key-tuple columns missing: {missing_key}. Available: {list(table.columns)}"
        )
    key_cols = KEY_TUPLE_COLUMN_NAMES
    missing_stat = [c for c in STATS_COLUMNS_REQUIRED if c not in cs]
    if missing_stat:
        raise KeyError(
            f"Stats columns missing: {missing_stat}. Available: {list(table.columns)}"
        )
    stat_cols = STATS_COLUMNS_REQUIRED.copy()
    if STATS_COLUMN_OPTIONAL in cs:
        stat_cols.append(STATS_COLUMN_OPTIONAL)
    return core_col, duration_col, key_cols, STATS_COLUMNS_REQUIRED, stat_cols


def _cell_repr(cell) -> str:
    """Human-readable cell value for logging."""
    if is_na(cell):
        return "<blank>"
    return repr(cell)


def key_tuple_field_values(row: Mapping[str, Any], key_cols: list[str]) -> dict[str, str]:
    """Map each key-tuple column name to the row's actual cell value (for diagnostics)."""
    return {col: _cell_repr(row.get(col)) for col in key_cols}


def parse_pad(cell) -> int | None:
    """Parse padded[logical] -> logical (int). Return None if not match."""
    if is_na(cell):
        return None
    s = str(cell).strip()
    m = PAD_PATTERN.match(s)
    if m is None:
        return None
    return int(m.group(2))


def _input1_blank(cell) -> bool:
    return is_na(cell) or (isinstance(cell, str) and not str(cell).strip())


def _all_input1_blank(row: Mapping[str, Any], input1_cols: list[str]) -> bool:
    return all(_input1_blank(row.get(c)) for c in input1_cols)


def _input1_cell_filled(
    col: str, row: Mapping[str, Any], pad_suffixes: tuple[str, ...]
) -> bool:
    """True if this INPUT_1 cell counts as non-blank (valid pad pattern or non-empty text)."""
    cell = row.get(col)
    is_pad = any(col.endswith(s) for s in pad_suffixes)
    if is_pad:
        return not _input1_blank(cell) and parse_pad(cell) is not None
    return not _input1_blank(cell)


def _all_input1_non_blank(
    row: Mapping[str, Any], input1_cols: list[str], pad_suffixes: tuple[str, ...]
) -> bool:
    return all(_input1_cell_filled(c, row, pad_suffixes) for c in input1_cols)


def _append_key_field(
    values: list,
    col: str,
    row: Mapping[str, Any],
    pad_suffixes: tuple[str, ...],
) -> str | None:
    """Append one key field to values. Returns failure_reason or None on success."""
    is_pad = any(col.endswith(s) for s in pad_suffixes)
    if is_pad:
        raw = row.get(col)
        val = parse_pad(raw)
        if val is None:
            return (
                f"PAD column {col!r} does not match padded[logical] pattern "
                f"(actual cell value: {_cell_repr(raw)})"
            )
        values.append(val)
        return None
    v = row.get(col)
    if is_na(v) or (isinstance(v, str) and not v.strip()):
        if col == "MATH FIDELITY":
            op_code = str(row.get("OP CODE", "")).strip().lower()
            mf_default = "HiFi4" if op_code in _MATH_FIDELITY_CALLER_CONTROLLED_OPS else MATH_FIDELITY_NA
            values.append(mf_default)
            return None
        return (
            f"blank or whitespace-only value in column {col!r} "
            f"(actual cell value: {_cell_repr(v)})"
        )
    if isinstance(v, (int, float)) and col == "OP CODE":
        values.append(str(int(v)) if v == v else str(v))
    else:
        values.append(v if isinstance(v, (int, float)) else str(v).strip())
    return None


def build_key_tuple(
    row: Mapping[str, Any],
    key_cols: list[str],
    pad_suffixes: tuple[str, ...],
) -> tuple[tuple | None, str | None]:
    """Build key tuple: 9, 16, or 23 fields (OP + INPUT_0 + MATH_FIDELITY; optional INPUT_1; optional INPUT_2).

    Rules: each of INPUT_1_* and INPUT_2_* is either all blank or all valid (no mixing).
    If INPUT_1_* is all blank, INPUT_2_* must be all blank (length 9).
    If INPUT_1_* is all set and INPUT_2_* all blank → length 16.
    If both slots are all set → length 23. INPUT_2 without INPUT_1 is invalid.

    Returns (tuple, None) on success, or (None, failure_reason) on skip.
    """
    if len(key_cols) != len(KEY_TUPLE_COLUMN_NAMES):
        raise ValueError(
            f"Expected {len(KEY_TUPLE_COLUMN_NAMES)} key columns, got {len(key_cols)}"
        )
    n0 = KEY_TUPLE_PREFIX_LEN
    n1 = KEY_TUPLE_INPUT_SLOT_LEN
    prefix_cols = key_cols[:n0]
    input1_cols = key_cols[n0 : n0 + n1]
    input2_cols = key_cols[n0 + n1 : n0 + n1 + n1]
    values: list = []
    for col in prefix_cols:
        err = _append_key_field(values, col, row, pad_suffixes)
        if err is not None:
            return None, err

    i1_blank = _all_input1_blank(row, input1_cols)
    i1_filled = _all_input1_non_blank(row, input1_cols, pad_suffixes)
    i2_blank = _all_input1_blank(row, input2_cols)
    i2_filled = _all_input1_non_blank(row, input2_cols, pad_suffixes)

    if i1_blank:
        if not i2_blank and not i2_filled:
            blank2 = [c for c in input2_cols if not _input1_cell_filled(c, row, pad_suffixes)]
            filled2 = [c for c in input2_cols if _input1_cell_filled(c, row, pad_suffixes)]
            return None, (
                "INPUT_2_* must be either all blank or all non-blank; "
                f"mixed row: non-blank in {filled2!r}, blank or invalid in {blank2!r}"
            )
        if i2_filled:
            return None, (
                "INPUT_2_* is set but INPUT_1_* is blank; ternary ops require both "
                "INPUT_1_* and INPUT_2_* when the third input is used."
            )
        return _normalize_math_fidelity(tuple(values)), None

    if not i1_filled:
        blank_cols = [c for c in input1_cols if not _input1_cell_filled(c, row, pad_suffixes)]
        filled_cols = [c for c in input1_cols if _input1_cell_filled(c, row, pad_suffixes)]
        return None, (
            "INPUT_1_* must be either all blank or all non-blank; "
            f"mixed row: non-blank in {filled_cols!r}, blank or invalid in {blank_cols!r}"
        )
    for col in input1_cols:
        err = _append_key_field(values, col, row, pad_suffixes)
        if err is not None:
            return None, err

    if i2_blank:
        return _normalize_math_fidelity(tuple(values)), None
    if not i2_filled:
        blank2 = [c for c in input2_cols if not _input1_cell_filled(c, row, pad_suffixes)]
        filled2 = [c for c in input2_cols if _input1_cell_filled(c, row, pad_suffixes)]
        return None, (
            "INPUT_2_* must be either all blank or all non-blank; "
            f"mixed row: non-blank in {filled2!r}, blank or invalid in {blank2!r}"
        )
    for col in input2_cols:
        err = _append_key_field(values, col, row, pad_suffixes)
        if err is not None:
            return None, err
    return _normalize_math_fidelity(tuple(values)), None


def _normalize_math_fidelity(key_t: tuple) -> tuple:
    """Replace math fidelity with ``N/A`` for ops where the caller doesn't control it."""
    if len(key_t) <= _MATH_FIDELITY_KEY_IDX:
        return key_t
    op_code = str(key_t[0]).strip().lower()
    if op_code not in _MATH_FIDELITY_CALLER_CONTROLLED_OPS:
        lst = list(key_t)
        lst[_MATH_FIDELITY_KEY_IDX] = MATH_FIDELITY_NA
        return tuple(lst)
    return key_t


def build_stats_row(
    row: Mapping[str, Any],
    dram_bw_gbps: float,
    stat_cols: list[str],
    duration_col: str,
    core_col: str,
) -> dict | None:
    """Build stats dict. None if non-numeric stat (skip row)."""
    duration_ns = row.get(duration_col)
    if duration_ns is None or is_na(duration_ns):
        return None
    try:
        duration_ns = float(duration_ns)
    except (TypeError, ValueError):
        return None
    util_pct = row.get(EXCEL_COL_DRAM_BW_UTIL)
    if util_pct is None or is_na(util_pct):
        util_pct = 0.0
    else:
        try:
            util_pct = float(util_pct)
        except (TypeError, ValueError):
            return None
    if not math.isfinite(util_pct):
        util_pct = 0.0
    # memory_traffic_bytes = (duration_ns/1e9) * (util/100) * DRAM_BW_Gbps * 1e9
    memory_traffic_bytes = (duration_ns / 1e9) * (util_pct / 100) * dram_bw_gbps * 1e9

    stats = {
        MASTER_DURATION_MS_KEY: duration_ns / NS_PER_MS,
        "memory_traffic": memory_traffic_bytes,
        OUTPUT_KEY_MEM_UTIL: util_pct,
        OUTPUT_KEY_NOC_UTIL: _stat_float(row.get(EXCEL_COL_NOC_UTIL)),
        OUTPUT_KEY_NOC_MULTICAST_UTIL: _stat_float(
            row.get(STATS_COLUMN_OPTIONAL, 0)
            if STATS_COLUMN_OPTIONAL in stat_cols
            else 0,
        ),
        OUTPUT_KEY_NPE_CONG_IMPACT_PCT: _stat_float(row.get(EXCEL_COL_NPE_CONG)),
        OUTPUT_KEY_VECTOR_PIPE_UTIL: _stat_float(row.get(EXCEL_COL_SFPU)),
        OUTPUT_KEY_MATRIX_PIPE_UTIL: _stat_float(row.get(EXCEL_COL_FPU)),
    }
    if frozenset(stats) != MASTER_SINGLE_STAT_KEYS_SET:
        raise RuntimeError(
            "build_stats_row keys must match MASTER_SINGLE_STAT_KEYS "
            f"(got {sorted(stats)!r} vs {list(MASTER_SINGLE_STAT_KEYS)!r})"
        )
    for k, v in stats.items():
        if v is None:
            return None
    return {k: stats[k] for k in MASTER_SINGLE_STAT_KEYS}


def _stat_float(x) -> float | None:
    """Parse a utilization-like cell; blanks → 0. NaN and non-finite floats → 0. Unparseable → None."""
    if is_na(x) or x == "" or (isinstance(x, str) and not x.strip()):
        return 0.0
    try:
        v = float(x)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(v):
        return 0.0
    return v


def r_squared(y: np.ndarray, y_pred: np.ndarray) -> float:
    ss_res = np.sum((y - y_pred) ** 2)
    ss_tot = np.sum((y - np.mean(y)) ** 2)
    if ss_tot == 0:
        return 0.0
    return 1.0 - (ss_res / ss_tot)


def _curve_y_all_nan_or_zero(y: np.ndarray) -> bool:
    """True if every sample is NaN or numeric zero (no curve fit; use identically-zero law)."""
    y = np.asarray(y, dtype=np.float64)
    if y.size == 0:
        return True
    return bool(np.all(np.isnan(y) | (y == 0)))


def _curve_zero_law_entry(stat_name: str) -> dict:
    """Curve-mode entry: stat is identically 0; a=b=0; R²=1 (exact under y≡0)."""
    return {
        "a": 0.0,
        "b": 0.0,
        "r2": 1.0,
        "equation": f"{stat_name} = 0",
    }


def fit_1d_linear(cores: np.ndarray, y: np.ndarray) -> tuple[np.ndarray, float]:
    """y = a*cores + b."""
    X = np.column_stack([cores, np.ones_like(cores)])
    coeffs, _, _, _ = np.linalg.lstsq(X, y, rcond=None)
    y_pred = X @ coeffs
    return coeffs, r_squared(y, y_pred)


def fit_1d_power(cores: np.ndarray, y: np.ndarray) -> tuple[np.ndarray, float]:
    """y = a * cores^b."""
    # Avoid zeros for log fit
    cores_safe = np.maximum(cores.astype(np.float64), 1e-9)
    y_safe = np.maximum(np.asarray(y, dtype=np.float64), 1e-9)

    def model(x: np.ndarray, a: float, b: float) -> np.ndarray:
        return a * (x ** b)

    try:
        a0 = np.nanmean(y_safe) / (np.nanmean(cores_safe) ** 0.5)
        (a, b), _ = curve_fit(
            model, cores_safe, y_safe, p0=[max(float(a0), 1e-6), -0.5], maxfev=10000
        )
        y_pred = model(cores_safe, a, b)
        return np.array([a, b]), r_squared(y_safe, y_pred)
    except Exception:
        return np.array([np.nanmean(y), 0.0]), 0.0


def choose_best_family_with_r2(
    cores: np.ndarray, duration: np.ndarray
) -> tuple[str, float, float]:
    """Return ('linear'|'power', R²_linear, R²_power) for duration vs core count."""
    _, r2_lin = fit_1d_linear(cores, duration)
    _, r2_pow = fit_1d_power(cores, duration)
    family = "linear" if r2_lin >= r2_pow else "power"
    return family, float(r2_lin), float(r2_pow)


def curve_entry_linear(
    cores: np.ndarray,
    coeffs: np.ndarray,
    stat_name: str,
    *,
    r2: float,
) -> dict:
    a, b = float(coeffs[0]), float(coeffs[1])
    return {
        "a": a,
        "b": b,
        "r2": float(r2),
        "equation": f"{stat_name} = {a:.6g} * Core_Count + {b:.6g}",
    }


def curve_entry_power(
    cores: np.ndarray,
    coeffs: np.ndarray,
    stat_name: str,
    *,
    r2: float,
) -> dict:
    a, b = float(coeffs[0]), float(coeffs[1])
    return {
        "a": a,
        "b": b,
        "r2": float(r2),
        "equation": f"{stat_name} = {a:.6g} * (Core_Count)^{b:.6g}",
    }


def _modified_z_outlier_mask(
    durations: np.ndarray,
    modified_z_threshold: float,
) -> tuple[np.ndarray, list[tuple[float, float]]]:
    """MAD-based modified Z (Iglewicz & Hoaglin). Returns (keep_mask, dropped (value_ms, |M*|)).

    No rejection when n < 3, MAD == 0, or any duration is non-finite.
    """
    d = np.asarray(durations, dtype=np.float64)
    n = len(d)
    if n < 3 or not np.all(np.isfinite(d)):
        return np.ones(n, dtype=bool), []
    med = float(np.median(d))
    mad = float(np.median(np.abs(d - med)))
    if mad == 0.0:
        return np.ones(n, dtype=bool), []
    mz = 0.6745 * (d - med) / mad
    keep = np.abs(mz) <= modified_z_threshold
    dropped = [(float(d[i]), float(abs(mz[i]))) for i in range(n) if not keep[i]]
    return keep, dropped


def group_and_sanitize(
    rows: list[tuple[tuple, float, dict, str]],
    duration_stat_key: str,
    cv_threshold: float,
) -> list[tuple[tuple, float, dict, str, list]]:
    """Group by (key_tuple, core_count). Drop groups with CV >= threshold.

    Before CV, duration outliers may be removed via modified Z (MAD); see
    CV_OUTLIER_MODIFIED_Z_THRESHOLD. Representative row is chosen from retained rows only.

    Returns accepted rows: (key_tuple, core_count, representative_stats, op_code, group_rows).
    """
    groups: dict[tuple[tuple, float], list[tuple[dict, float, str]]] = defaultdict(list)
    for key_t, core_count, stats, op_code in rows:
        groups[(key_t, core_count)].append((stats, stats[duration_stat_key], op_code))

    accepted = []
    for (key_t, core_count), group_data in groups.items():
        durations = np.array([d for _, d, _ in group_data], dtype=np.float64)
        working_group = group_data
        working_durations = durations

        if len(durations) >= 2:
            keep, dropped_pairs = _modified_z_outlier_mask(
                durations, CV_OUTLIER_MODIFIED_Z_THRESHOLD
            )
            n_kept = int(np.sum(keep))
            if dropped_pairs:
                dropped_str = ", ".join(
                    f"{val:.6g} ms (|M*|={mz:.3g})" for val, mz in dropped_pairs
                )
                kept_vals = sorted(float(x) for x in durations[keep])
                kept_str = ", ".join(f"{v:.6g} ms" for v in kept_vals)
                if n_kept >= 2:
                    logger.info(
                        "CV gate: ignoring {} duration outlier(s) (modified |M*| > {}): {}; "
                        "accepted for CV (n={}): {}; core_count={}; key_tuple={}",
                        len(dropped_pairs),
                        CV_OUTLIER_MODIFIED_Z_THRESHOLD,
                        dropped_str,
                        n_kept,
                        kept_str,
                        core_count,
                        list(key_t),
                    )
                else:
                    logger.info(
                        "CV gate: ignoring {} duration outlier(s) (modified |M*| > {}): {}; "
                        "would retain only n={} ({}); reverting to full group for CV. "
                        "core_count={}; key_tuple={}",
                        len(dropped_pairs),
                        CV_OUTLIER_MODIFIED_Z_THRESHOLD,
                        dropped_str,
                        n_kept,
                        kept_str if kept_str else "(none)",
                        core_count,
                        list(key_t),
                    )
            if n_kept >= 2:
                working_group = [g for g, k in zip(group_data, keep) if k]
                working_durations = np.array([d for _, d, _ in working_group])

        if len(working_durations) < 2:
            cv = 0.0
        else:
            mean_d = float(np.mean(working_durations))
            std_d = float(np.std(working_durations))
            cv = (std_d / mean_d) if mean_d != 0 else float("inf")
        if cv > cv_threshold:
            logger.warning(
                "Skipping group (CV={cv:.2%}): core_count={cc}; key_tuple={kt}",
                cv=cv,
                cc=core_count,
                kt=list(key_t),
            )
            continue
        # Representative: median duration among retained rows only
        sorted_group = sorted(working_group, key=lambda x: x[1])
        idx = (len(sorted_group) - 1) // 2
        rep_stats = sorted_group[idx][0]
        op_code = sorted_group[idx][2]
        accepted.append(
            (key_t, core_count, rep_stats, op_code, [g[0] for g in working_group])
        )
    return accepted


def build_master_dict(
    accepted: list[tuple[tuple, float, dict, str, list]],
    mode: str,
) -> tuple[dict[tuple, dict], dict[tuple, dict]]:
    """Build master dict from grouped rows.

    ``mode`` is ``"single"`` (each ``--model-run``) or ``"curve"`` (each ``--sweep-run``) — not a CLI flag.

    * ``"single"``: one distinct core count per key; value is ``MASTER_ENTRY_TYPE_SINGLE``,
      ``MASTER_SINGLE_NUM_CORES_KEY``, and the stats dict fields (flat; includes matmul).
    * ``"curve"``: at least two core counts; matmul only; value is ``MASTER_ENTRY_TYPE_CURVE``,
      ``MASTER_CURVE_FAMILY_KEY``, and per-stat fit dicts.

    Stats that are all NaN or 0 across cores get a=0, b=0, equation ``... = 0``, r2=1.

    Returns ``(master, curve_meta)``. ``curve_meta`` is non-empty only for ``mode="curve"`` (fit summaries).
    """
    # By key_tuple: list of (core_count, rep_stats, op_code, group_stats_list)
    by_key: dict[tuple, list[tuple[float, dict, str, list]]] = defaultdict(list)
    for key_t, core_count, rep_stats, op_code, group_rows in accepted:
        by_key[key_t].append((core_count, rep_stats, op_code, group_rows))

    master: dict[tuple, dict] = {}
    curve_meta: dict[tuple, dict] = {}
    for key_t, entries in by_key.items():
        core_counts = sorted(set(e[0] for e in entries))
        n_cores = len(core_counts)
        if mode == "single":
            if n_cores != 1:
                cc_counts: defaultdict[float, int] = defaultdict(int)
                for e in entries:
                    cc_counts[e[0]] += 1
                cc = float(max(cc_counts, key=cc_counts.get))  # type: ignore[arg-type]
                logger.warning(
                    "key_tuple has {} distinct CORE COUNTs {} in single mode; "
                    "using most frequent ({}, {} of {} rows); key_tuple={}",
                    n_cores,
                    core_counts,
                    int(cc),
                    cc_counts[cc],
                    sum(cc_counts.values()),
                    list(key_t),
                )
                entries = [e for e in entries if float(e[0]) == cc]
            else:
                cc = float(core_counts[0])
            core_key = _single_mode_core_count_key(cc, key_t)
            rep = next(e[1] for e in entries if float(e[0]) == cc)
            master[key_t] = {
                MASTER_ENTRY_TYPE_KEY: MASTER_ENTRY_TYPE_SINGLE,
                MASTER_SINGLE_NUM_CORES_KEY: core_key,
                **rep,
            }
        else:  # curve
            if n_cores < 2:
                continue
            op_code = entries[0][2]
            if op_code != POLARIS_LAYER_MATMUL:
                logger.error(
                    "key_tuple with multiple CORE COUNT has OP CODE {!r}, expected {!r}; key_tuple={}",
                    op_code,
                    POLARIS_LAYER_MATMUL,
                    list(key_t),
                )
                sys.exit(1)
            cores_arr = np.array(core_counts, dtype=np.float64)
            # Collect stat vectors across group rows for this key (one value per core_count: use rep)
            rep_by_core = {e[0]: e[1] for e in entries}
            stat_names = list(rep_by_core[core_counts[0]].keys())
            dur_arr = np.array(
                [rep_by_core[c][MASTER_DURATION_MS_KEY] for c in core_counts],
                dtype=np.float64,
            )
            if _curve_y_all_nan_or_zero(dur_arr):
                family, r2_lin_d, r2_pow_d = "linear", 0.0, 0.0
            else:
                family, r2_lin_d, r2_pow_d = choose_best_family_with_r2(
                    cores_arr, dur_arr
                )
            curve_val = {}
            for stat_name in stat_names:
                y_arr = np.array(
                    [rep_by_core[c][stat_name] for c in core_counts],
                    dtype=np.float64,
                )
                if _curve_y_all_nan_or_zero(y_arr):
                    curve_val[stat_name] = _curve_zero_law_entry(stat_name)
                    continue
                try:
                    if family == "linear":
                        coeffs, r2_stat = fit_1d_linear(cores_arr, y_arr)
                        curve_val[stat_name] = curve_entry_linear(
                            cores_arr,
                            coeffs,
                            stat_name,
                            r2=r2_stat,
                        )
                    else:
                        coeffs, r2_stat = fit_1d_power(cores_arr, y_arr)
                        curve_val[stat_name] = curve_entry_power(
                            cores_arr,
                            coeffs,
                            stat_name,
                            r2=r2_stat,
                        )
                except Exception:
                    med = float(np.median(y_arr))
                    logger.warning(
                        "Curve fit failed for stat {}, using constant {}; key_tuple={}",
                        repr(stat_name),
                        med,
                        list(key_t),
                    )
                    y_const = np.full_like(y_arr, med, dtype=np.float64)
                    r2_const = r_squared(y_arr, y_const)
                    curve_val[stat_name] = {
                        "a": med,
                        "b": 0.0,
                        "r2": float(r2_const),
                        "equation": f"{stat_name} = {med:.6g}",
                    }
            master[key_t] = {
                MASTER_ENTRY_TYPE_KEY: MASTER_ENTRY_TYPE_CURVE,
                MASTER_CURVE_FAMILY_KEY: family,
                **curve_val,
            }
            curve_meta[key_t] = {
                "family": family,
                "r2_linear_duration": r2_lin_d,
                "r2_power_duration": r2_pow_d,
                "core_counts": [float(c) for c in core_counts],
            }
    return master, curve_meta


def is_single_entry(val: dict) -> bool:
    """True iff ``entry_type`` is ``single`` (YAML loads require this field)."""
    return (
        isinstance(val, dict)
        and val.get(MASTER_ENTRY_TYPE_KEY) == MASTER_ENTRY_TYPE_SINGLE
    )


def is_curve_entry(val: dict) -> bool:
    """True iff ``entry_type`` is ``curve`` (YAML loads require this field)."""
    return (
        isinstance(val, dict)
        and val.get(MASTER_ENTRY_TYPE_KEY) == MASTER_ENTRY_TYPE_CURVE
    )


def is_hybrid_entry(val: dict) -> bool:
    """True iff ``entry_type`` is ``hybrid`` (matmul: nested ``single`` / ``curve``)."""
    return (
        isinstance(val, dict)
        and val.get(MASTER_ENTRY_TYPE_KEY) == MASTER_ENTRY_TYPE_HYBRID
    )


def _matmul_entry_to_hybrid(val: dict) -> dict:
    """Convert matmul ``single`` / ``curve`` / ``hybrid`` payload to a hybrid dict."""
    if is_hybrid_entry(val):
        out: dict[str, Any] = {MASTER_ENTRY_TYPE_KEY: MASTER_ENTRY_TYPE_HYBRID}
        sk = val.get(MASTER_HYBRID_SINGLE_KEY)
        if sk:
            out[MASTER_HYBRID_SINGLE_KEY] = normalize_flat_single_payload(dict(sk))
        ck = val.get(MASTER_HYBRID_CURVE_KEY)
        if ck:
            out[MASTER_HYBRID_CURVE_KEY] = dict(ck)
        if MASTER_HYBRID_SINGLE_KEY not in out and MASTER_HYBRID_CURVE_KEY not in out:
            raise ValueError("hybrid entry has neither single nor curve branch")
        return out
    if is_single_entry(val):
        flat = {
            k: v for k, v in val.items() if k != MASTER_ENTRY_TYPE_KEY
        }
        return {
            MASTER_ENTRY_TYPE_KEY: MASTER_ENTRY_TYPE_HYBRID,
            MASTER_HYBRID_SINGLE_KEY: normalize_flat_single_payload(flat),
        }
    if is_curve_entry(val):
        curve_inner = {
            k: v for k, v in val.items() if k != MASTER_ENTRY_TYPE_KEY
        }
        return {
            MASTER_ENTRY_TYPE_KEY: MASTER_ENTRY_TYPE_HYBRID,
            MASTER_HYBRID_CURVE_KEY: curve_inner,
        }
    t = val.get(MASTER_ENTRY_TYPE_KEY)
    raise ValueError(
        f"Matmul merge expected single, curve, or hybrid entry, got {t!r}"
    )


def _merge_hybrid_matmul(old: dict, new: dict) -> dict:
    """Merge two hybrid matmul entries; later ``single`` / ``curve`` branches win."""
    os_single = old.get(MASTER_HYBRID_SINGLE_KEY)
    ns_single = new.get(MASTER_HYBRID_SINGLE_KEY)
    if ns_single is not None:
        merged_single = normalize_flat_single_payload(
            {**(dict(os_single) if os_single else {}), **dict(ns_single)}
        )
    elif os_single is not None:
        merged_single = normalize_flat_single_payload(dict(os_single))
    else:
        merged_single = None
    if merged_single is not None and len(merged_single) == 0:
        merged_single = None
    oc = old.get(MASTER_HYBRID_CURVE_KEY)
    nc = new.get(MASTER_HYBRID_CURVE_KEY)
    merged_curve = dict(nc) if nc is not None else (dict(oc) if oc is not None else None)
    out: dict = {MASTER_ENTRY_TYPE_KEY: MASTER_ENTRY_TYPE_HYBRID}
    if merged_single:
        out[MASTER_HYBRID_SINGLE_KEY] = merged_single
    if merged_curve is not None:
        out[MASTER_HYBRID_CURVE_KEY] = merged_curve
    if MASTER_HYBRID_SINGLE_KEY not in out and MASTER_HYBRID_CURVE_KEY not in out:
        raise ValueError("merged matmul hybrid would be empty")
    return out


def merge_entry_for_key(
    key_t: tuple,
    old_val: dict | None,
    new_val: dict,
) -> dict:
    """Merge one new entry into an optional existing entry for the same key tuple."""
    op = str(key_t[0]) if key_t else ""
    if op != POLARIS_LAYER_MATMUL:
        if old_val is None:
            return dict(new_val)
        old_kind = _merge_entry_kind(old_val)
        new_kind = _merge_entry_kind(new_val)
        if old_kind != new_kind:
            raise ValueError(
                f"Key {tuple_to_labeled_key_map(key_t)} has existing entry type "
                f"{old_kind!r} but this run produced {new_kind!r}. Exit."
            )
        return dict(new_val)
    h_new = _matmul_entry_to_hybrid(new_val)
    if old_val is None:
        return h_new
    h_old = _matmul_entry_to_hybrid(old_val)
    return _merge_hybrid_matmul(h_old, h_new)


def _merge_entry_kind(val: dict) -> str:
    """Return ``single`` or ``curve``; raise if ``entry_type`` is missing or invalid."""
    if is_single_entry(val):
        return MASTER_ENTRY_TYPE_SINGLE
    if is_curve_entry(val):
        return MASTER_ENTRY_TYPE_CURVE
    t = val.get(MASTER_ENTRY_TYPE_KEY) if isinstance(val, dict) else None
    raise ValueError(
        f"Master entry requires {MASTER_ENTRY_TYPE_KEY!r} "
        f"{MASTER_ENTRY_TYPE_SINGLE!r}, {MASTER_ENTRY_TYPE_CURVE!r}, or "
        f"{MASTER_ENTRY_TYPE_HYBRID!r} (matmul), got {t!r}"
    )


def merge_master(
    existing: dict[tuple, dict],
    new: dict[tuple, dict],
    emit_overwrite_warning: bool = True,
) -> dict[tuple, dict]:
    """Merge new into existing. Non-matmul: same key + same entry_type -> overwrite.
    Matmul: ``single`` and ``curve`` combine into ``hybrid``. Same key + different
    non-matmul entry types -> raise.

    Dry-run accounting of which keys were added vs updated is done in ``main()`` from
    the on-disk key set and the union of keys present in each input Excel.
    """
    result = dict(existing)
    for key_t, new_val in new.items():
        had = key_t in result
        if had and emit_overwrite_warning:
            logger.warning(
                "Entry already exists for key_tuple, overwriting: {}",
                tuple_to_labeled_key_map(key_t),
            )
        old_val = result.get(key_t)
        try:
            merged_val = merge_entry_for_key(key_t, old_val, new_val)
        except ValueError:
            raise
        result[key_t] = merged_val
    return result


def canonicalize_master_for_write(master: dict[tuple, dict]) -> dict[tuple, dict]:
    """Ensure matmul keys use ``hybrid`` on disk (plain single/curve from Excel builds)."""
    out: dict[tuple, dict] = {}
    for key_t, val in master.items():
        op = str(key_t[0]) if key_t else ""
        if op == POLARIS_LAYER_MATMUL and (
            is_single_entry(val) or is_curve_entry(val)
        ):
            out[key_t] = _matmul_entry_to_hybrid(val)
        else:
            out[key_t] = dict(val)
    return out


def _curve_entry_for_dry_run(val: dict) -> dict | None:
    """Synthetic curve entry for reporting, or ``None``."""
    if is_curve_entry(val):
        return val
    if is_hybrid_entry(val):
        inner = val.get(MASTER_HYBRID_CURVE_KEY)
        if not inner:
            return None
        return {MASTER_ENTRY_TYPE_KEY: MASTER_ENTRY_TYPE_CURVE, **dict(inner)}
    return None


def _report_tuple_sort_key(key_t: tuple) -> tuple:
    """Lexicographic order for key tuples with mixed str/int/float (avoids Py3 str/int compare errors)."""
    out: list = []
    for x in key_t:
        if isinstance(x, str):
            out.append((0, x))
        elif isinstance(x, (bool, np.bool_)):
            out.append((1, bool(x)))
        elif isinstance(x, (int, np.integer)):
            out.append((2, int(x)))
        elif isinstance(x, (float, np.floating)):
            out.append((3, float(x)))
        else:
            out.append((4, str(x)))
    return tuple(out)


def _curve_dry_run_lines(
    key_t: tuple,
    entry: dict,
    meta: dict | None,
) -> list[str]:
    """Indented lines: fit summary + per-stat a, b, equation (curve entries only)."""
    lines: list[str] = []
    m = meta or {}
    fam = entry.get(MASTER_CURVE_FAMILY_KEY) or m.get("family", "?")
    cores = m.get("core_counts", [])
    rlin = m.get("r2_linear_duration")
    rpow = m.get("r2_power_duration")
    rlin_s = f"{rlin:.6g}" if isinstance(rlin, (int, float)) and rlin == rlin else "?"
    rpow_s = f"{rpow:.6g}" if isinstance(rpow, (int, float)) and rpow == rpow else "?"
    lines.append(
        f"      curve: family={fam!r} (all stats); core_counts={cores}; "
        f"duration R² linear={rlin_s}, power={rpow_s}"
    )
    stat_keys = sorted(
        k
        for k, v in entry.items()
        if k != MASTER_CURVE_FAMILY_KEY
        and isinstance(v, dict)
        and ("equation" in v or "a" in v)
    )
    for stat_name in stat_keys:
        sub = entry[stat_name]
        a = sub.get("a")
        b = sub.get("b")
        eq = sub.get("equation", "")
        r2 = sub.get("r2")
        r2_s = f"{r2:.6g}" if isinstance(r2, (int, float)) and r2 == r2 else "?"
        lines.append(f"      {stat_name}: a={a!s}, b={b!s}; R²={r2_s}")
        lines.append(f"        equation: {eq}")
    return lines


def format_dry_run_report(
    out_path: Path,
    keys_added: list[tuple],
    keys_updated: list[tuple],
    *,
    merged: dict[tuple, dict],
    curve_meta: dict[tuple, dict] | None = None,
) -> str:
    """Human-readable report of what would be written (no file write).

    ``keys_added`` / ``keys_updated`` are logical key tuples that appear in any
    input Excel for this run: added = not present in the on-disk master before
    the run; updated = already on disk. Each key appears at most once per list
    (no duplicates from multiple ``--model-run`` / ``--sweep-run`` arguments).
    """
    added_sorted = sorted(keys_added, key=_report_tuple_sort_key)
    updated_sorted = sorted(keys_updated, key=_report_tuple_sort_key)
    cm = curve_meta or {}
    lines = [
        "Dry run: --update not set; output file will not be modified.",
        f"Output path: {out_path}",
        f"Would add {len(keys_added)} entr{'y' if len(keys_added) == 1 else 'ies'}.",
    ]
    for k in added_sorted:
        lines.append(f"  + {tuple_to_labeled_key_map(k)}")
        ce = _curve_entry_for_dry_run(merged.get(k, {}))
        if ce:
            lines.extend(_curve_dry_run_lines(k, ce, cm.get(k)))
    lines.append(
        f"Would update {len(keys_updated)} entr{'y' if len(keys_updated) == 1 else 'ies'}."
    )
    for k in updated_sorted:
        lines.append(f"  ~ {tuple_to_labeled_key_map(k)}")
        ce = _curve_entry_for_dry_run(merged.get(k, {}))
        if ce:
            lines.extend(_curve_dry_run_lines(k, ce, cm.get(k)))
    return "\n".join(lines)


def serialize_master_for_yaml(master: dict[tuple, dict]) -> dict:
    """Build YAML document: schema_name, schema_version, entries.

    Each list item is ``{key: labeled_key_map, value: entry_dict}`` (see wire constants).
    """
    return {
        MASTER_YAML_SCHEMA_NAME_KEY: MASTER_YAML_SCHEMA_NAME,
        MASTER_YAML_SCHEMA_VERSION_KEY: MASTER_YAML_SCHEMA_VERSION,
        MASTER_YAML_ENTRIES_KEY: [
            {
                MASTER_YAML_RECORD_KEY_FIELD: tuple_to_labeled_key_map(k),
                MASTER_YAML_ENTRY_VALUE_FIELD: v,
            }
            for k, v in master.items()
        ],
    }


def process_excel_to_master(
    input_spec: str,
    mode: str,
    dram_bw_gbps: float,
) -> tuple[int, dict[tuple, dict], dict[tuple, dict], Path]:
    """Load one Excel or CSV input spec and build master dict.

    Returns ``(exit_code, master, curve_meta, excel_path)``. ``exit_code`` is ``0`` on success,
    ``1`` for fatal sheet row errors / empty data, ``2`` for missing file / load errors.
    On failure, ``master`` and ``curve_meta`` are empty dicts.
    """
    sheet_name, file_path = parse_input_spec(input_spec)
    excel_path = Path(file_path)
    if not excel_path.exists():
        logger.error("Input file not found: {}", excel_path)
        return 2, {}, {}, excel_path

    try:
        table = load_input_table(sheet_name, excel_path)
    except (KeyError, ValueError, OSError, UnicodeDecodeError, csv.Error, BadZipFile) as e:
        logger.error("{}", e)
        return 2, {}, {}, excel_path

    apply_merged_ops_stat_aliases(table)

    original_opcodes = [row.get("OP CODE", "") for row in table.rows]
    apply_polaris_layer_type_column(table)

    other_ops: dict[str, int] = {}
    for orig, row in zip(original_opcodes, table.rows):
        if row.get("OP CODE") == "other":
            key = str(orig).strip() if orig else "<blank>"
            other_ops[key] = other_ops.get(key, 0) + 1
    if other_ops:
        detail = "; ".join(f"{name!r} ({n} row{'s' if n > 1 else ''})" for name, n in sorted(other_ops.items()))
        logger.error(
            "Refusing to build LUT: {} profiler OP CODE(s) mapped to 'other' "
            "(unrecognized by op_canonical.py): {}. "
            "Add prefix rules for these ops in tools/profiling/op_canonical.py "
            "before re-running.",
            len(other_ops),
            detail,
        )
        return 1, {}, {}, excel_path

    try:
        core_col, duration_col, key_cols, stat_required, stat_cols = validate_columns(
            table
        )
    except KeyError as e:
        logger.error("{}", e)
        return 2, {}, {}, excel_path

    pad_suffixes = (
        "_W_PAD[LOGICAL]",
        "_Z_PAD[LOGICAL]",
        "_Y_PAD[LOGICAL]",
        "_X_PAD[LOGICAL]",
    )

    rows: list[tuple[tuple, float, dict, str]] = []
    for idx, row in enumerate(table.rows):
        key_t, key_fail = build_key_tuple(row, key_cols, pad_suffixes)
        if key_t is None:
            logger.warning(
                "Skipping row (invalid key tuple) index={}: {}; key_tuple_column_values={}",
                idx,
                key_fail,
                key_tuple_field_values(row, key_cols),
            )
            continue
        if is_na(row.get(duration_col)) or (
            isinstance(row.get(duration_col), str)
            and not str(row.get(duration_col)).strip()
        ):
            logger.error(
                "{} blank at row {}; key_tuple_column_values={}",
                EXCEL_DURATION_COLUMN,
                idx,
                key_tuple_field_values(row, key_cols),
            )
            return 1, {}, {}, excel_path
        stats = build_stats_row(
            row, dram_bw_gbps, stat_cols, duration_col, core_col
        )
        if stats is None:
            logger.warning(
                "Skipping row (non-numeric stat) index={}; key_tuple={}; key_tuple_column_values={}",
                idx,
                list(key_t),
                key_tuple_field_values(row, key_cols),
            )
            continue
        cc_raw = row.get(core_col)
        if cc_raw is None or is_na(cc_raw):
            logger.warning(
                "Skipping row (missing CORE COUNT) index={}; key_tuple={}",
                idx,
                list(key_t),
            )
            continue
        try:
            core_count = float(cc_raw)
        except (TypeError, ValueError):
            logger.warning(
                "Skipping row (invalid CORE COUNT) index={}; CORE_COUNT_cell={}; key_tuple={}",
                idx,
                _cell_repr(row.get(core_col)),
                list(key_t),
            )
            continue
        op_code = key_t[0] if isinstance(key_t[0], str) else str(key_t[0])
        rows.append((key_t, core_count, stats, op_code))

    if not rows:
        logger.error("No rows after parsing.")
        return 1, {}, {}, excel_path

    accepted = group_and_sanitize(rows, MASTER_DURATION_MS_KEY, CV_THRESHOLD)
    master, curve_meta = build_master_dict(accepted, mode)
    return 0, master, curve_meta, excel_path


def _default_output_path(args: argparse.Namespace) -> Path:
    if args.output is not None:
        return args.output
    first = args.model_run[0] if args.model_run else args.sweep_run[0]
    _, fp = parse_input_spec(first)
    p = Path(fp)
    return p.parent / f"{p.stem}_master.yaml"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build master performance map from Excel or CSV; write or update YAML."
    )
    parser.add_argument(
        "--model-run",
        action="append",
        default=[],
        metavar="SHEET@FILE|FILE",
        help=(
            "Excel (.xlsx/.xlsm) or CSV from a full model run (single mode: one CORE COUNT per key). "
            "CSV is read with stdlib csv. Repeatable."
        ),
    )
    parser.add_argument(
        "--sweep-run",
        action="append",
        default=[],
        metavar="SHEET@FILE|FILE",
        help=(
            "Excel or CSV from a core sweep (curve mode for matmul). CSV is read with stdlib csv. Repeatable."
        ),
    )
    parser.add_argument(
        "--dram-bw-gbps",
        required=True,
        type=float,
        help="DRAM bandwidth in GB/s for memory_traffic (bytes).",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=None,
        help="Output YAML path. Default: {first_input_stem}_master.yaml next to the first input file.",
    )
    parser.add_argument(
        "--update",
        action="store_true",
        help=(
            "Write merged master to the output YAML file. If omitted, only report what "
            "would be added/updated (no file write)."
        ),
    )
    args = parser.parse_args()
    if not args.model_run and not args.sweep_run:
        parser.error("Provide at least one --model-run and/or --sweep-run.")
    return args


def _load_existing_master_for_merge(out_path: Path) -> tuple[int, dict[tuple, dict]]:
    yaml_path = out_path.resolve()
    try:
        return 0, load_existing_yaml(out_path)
    except (ValueError, TypeError) as e:
        logger.error(
            "Existing output file is not a valid tt_perf master YAML ({}): {}",
            yaml_path,
            e,
        )
        logger.error(
            "Required shape: top-level mapping with {}={!r}, {}={!r}, and {} "
            "as a list of objects; each object must have field {!r} (labeled record key) "
            "and field {!r} (entry: single, curve, or hybrid for matmul). "
            "Record keys: tt_perf_master_schema.KEY_TUPLE_YAML_KEYS. "
            "Reference: doc/YAML_MASTER_FORMAT.md.",
            MASTER_YAML_SCHEMA_NAME_KEY,
            MASTER_YAML_SCHEMA_NAME,
            MASTER_YAML_SCHEMA_VERSION_KEY,
            MASTER_YAML_SCHEMA_VERSION,
            MASTER_YAML_ENTRIES_KEY,
            MASTER_YAML_RECORD_KEY_FIELD,
            MASTER_YAML_ENTRY_VALUE_FIELD,
        )
        return 2, {}
    except yaml.YAMLError as e:
        logger.error("YAML parse error in {}: {}", yaml_path, e)
        return 2, {}
    except OSError as e:
        logger.error("Cannot read existing output file {}: {}", yaml_path, e)
        return 2, {}
    except Exception as e:
        logger.error("Unexpected error loading {}: {}", yaml_path, e)
        return 2, {}


def main() -> int:
    args = parse_args()
    curve_meta: dict[tuple, dict] = {}

    out_path = _default_output_path(args)
    merged: dict[tuple, dict] = {}
    if out_path.exists():
        ycode, existing = _load_existing_master_for_merge(out_path)
        if ycode != 0:
            return ycode
        merged = existing

    initial_keys = frozenset(merged.keys())
    keys_from_inputs: set[tuple] = set()

    for spec in args.model_run:
        code, master, cm, _ = process_excel_to_master(
            spec, "single", args.dram_bw_gbps
        )
        if code != 0:
            return code
        keys_from_inputs |= set(master.keys())
        try:
            merged = merge_master(
                merged,
                master,
                emit_overwrite_warning=args.update,
            )
        except ValueError as e:
            logger.error("{}", e)
            return 2

    for spec in args.sweep_run:
        code, master, cm, _ = process_excel_to_master(
            spec, "curve", args.dram_bw_gbps
        )
        if code != 0:
            return code
        curve_meta.update(cm)
        keys_from_inputs |= set(master.keys())
        try:
            merged = merge_master(
                merged,
                master,
                emit_overwrite_warning=args.update,
            )
        except ValueError as e:
            logger.error("{}", e)
            return 2

    merged = canonicalize_master_for_write(merged)

    keys_added = sorted(
        keys_from_inputs - initial_keys, key=_report_tuple_sort_key
    )
    keys_updated = sorted(
        keys_from_inputs & initial_keys, key=_report_tuple_sort_key
    )

    if not args.update:
        logger.info(
            "{}",
            format_dry_run_report(
                out_path,
                keys_added,
                keys_updated,
                merged=merged,
                curve_meta=curve_meta,
            ),
        )
        return 0

    payload = serialize_master_for_yaml(merged)
    with open(out_path, "w") as f:
        yaml.dump(
            payload,
            f,
            default_flow_style=False,
            sort_keys=False,
            allow_unicode=True,
        )
    logger.info("Wrote master to {}", out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
