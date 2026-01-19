# SPDX-FileCopyrightText: (C) 2025 Tenstorrent AI ULC
# SPDX-License-Identifier: Apache-2.0
"""
Unit tests for ttsi_corr.excel_writer module.

Tests the ExcelFormatter class and write_csv function to ensure proper
Excel formatting utilities functionality.
"""

import pytest
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from tools.ttsi_corr.excel_writer import ExcelFormatter, write_csv


class TestExcelFormatter:
    """Tests for the ExcelFormatter class."""
    
    def test_col_letter_single_letters(self):
        """Test column letter conversion for single-letter columns (A-Z)."""
        formatter = ExcelFormatter()
        
        assert formatter.col_letter(1) == 'A'
        assert formatter.col_letter(2) == 'B'
        assert formatter.col_letter(26) == 'Z'
    
    def test_col_letter_double_letters(self):
        """Test column letter conversion for double-letter columns (AA-ZZ)."""
        formatter = ExcelFormatter()
        
        assert formatter.col_letter(27) == 'AA'
        assert formatter.col_letter(28) == 'AB'
        assert formatter.col_letter(52) == 'AZ'
        assert formatter.col_letter(53) == 'BA'
        assert formatter.col_letter(702) == 'ZZ'
    
    def test_col_letter_triple_letters(self):
        """Test column letter conversion for triple-letter columns (AAA+)."""
        formatter = ExcelFormatter()
        
        assert formatter.col_letter(703) == 'AAA'
        assert formatter.col_letter(704) == 'AAB'
    
    def test_col_letter_static_method(self):
        """Test that col_letter can be called as a static method."""
        # Should work without instantiating the class
        assert ExcelFormatter.col_letter(1) == 'A'
        assert ExcelFormatter.col_letter(27) == 'AA'
    
    def test_apply_number_formats(self):
        """Test apply_number_formats method."""
        formatter = ExcelFormatter()
        wb = Workbook()
        ws = wb.active
        
        # Sample data
        headers = ['Workload', 'Batch', 'Score', 'Ratio']
        comparison = [
            {'Workload': 'BERT', 'Batch': 16, 'Score': 100.5, 'Ratio': 0.95},
            {'Workload': 'ResNet', 'Batch': 32, 'Score': 200.75, 'Ratio': 1.05}
        ]
        
        # Write headers and data
        ws.append(headers)
        for row in comparison:
            ws.append(list(row.values()))
        
        # Apply formatting
        formatter.apply_number_formats(ws, headers, comparison)
        
        # Check Batch column (column 2) has integer format
        assert ws['B2'].number_format == '0'
        assert ws['B3'].number_format == '0'
        
        # Check Score column (column 3) has comma format
        assert ws['C2'].number_format == '#,##0.00'
        assert ws['C3'].number_format == '#,##0.00'
        
        # Check Ratio column (column 4) has comma format
        assert ws['D2'].number_format == '#,##0.00'
        assert ws['D3'].number_format == '#,##0.00'
    
    def test_apply_borders(self):
        """Test apply_borders method."""
        formatter = ExcelFormatter()
        wb = Workbook()
        ws = wb.active
        
        # Write some data
        for i in range(3):
            ws.append([f'Cell{i}{j}' for j in range(4)])
        
        # Apply borders
        formatter.apply_borders(ws, num_rows=3, num_cols=4)
        
        # Check that all cells have borders
        for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=4):
            for cell in row:
                assert cell.border.left.style == 'thin'
                assert cell.border.right.style == 'thin'
                assert cell.border.top.style == 'thin'
                assert cell.border.bottom.style == 'thin'
    
    def test_apply_freeze_panes(self):
        """Test apply_freeze_panes method."""
        formatter = ExcelFormatter()
        wb = Workbook()
        ws = wb.active
        
        # Apply freeze panes
        formatter.apply_freeze_panes(ws, freeze_col='E', freeze_row=2)
        
        # Check freeze panes is set correctly
        assert ws.freeze_panes == 'E2'


class TestWriteCsv:
    """Tests for the write_csv function."""
    
    def test_write_csv_basic(self, tmp_path):
        """Test basic CSV writing functionality."""
        output_file = tmp_path / 'test_output.csv'
        
        comparison = [
            {'Workload': 'BERT', 'Score': 100.0, 'Ratio': 0.95},
            {'Workload': 'ResNet', 'Score': 200.0, 'Ratio': 1.05}
        ]
        
        write_csv(comparison, output_file)
        
        # Check file was created
        assert output_file.exists()
        
        # Read back and verify content
        with open(output_file, 'r') as f:
            lines = f.readlines()
        
        # Check header
        assert 'Workload' in lines[0]
        assert 'Score' in lines[0]
        assert 'Ratio' in lines[0]
        
        # Check data rows
        assert 'BERT' in lines[1]
        assert 'ResNet' in lines[2]
    
    def test_write_csv_empty_raises_error(self, tmp_path):
        """Test that write_csv raises ValueError for empty list."""
        output_file = tmp_path / 'test_output.csv'
        
        with pytest.raises(ValueError, match='Comparison list cannot be empty'):
            write_csv([], output_file)
    
    def test_write_csv_preserves_column_order(self, tmp_path):
        """Test that CSV preserves the column order from input dictionaries."""
        output_file = tmp_path / 'test_output.csv'
        
        # Use ordered dict to ensure column order
        comparison = [
            {'Col1': 'A', 'Col2': 'B', 'Col3': 'C'},
            {'Col1': 'D', 'Col2': 'E', 'Col3': 'F'}
        ]
        
        write_csv(comparison, output_file)
        
        # Read back and verify column order
        with open(output_file, 'r') as f:
            header = f.readline().strip()
        
        assert header == 'Col1,Col2,Col3'

