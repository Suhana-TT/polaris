# SPDX-FileCopyrightText: (C) 2025 Tenstorrent AI ULC
# SPDX-License-Identifier: Apache-2.0
"""
Unit tests for ttsi_corr.chart_builder module.

Tests the ScurveChartBuilder class and add_scurve_chart function to ensure
proper S-curve chart generation functionality.
"""

import math
import pytest
from openpyxl import Workbook  # type: ignore[import-untyped]

from tools.ttsi_corr.chart_builder import ScurveChartBuilder, add_scurve_chart


class TestScurveChartBuilder:
    """Tests for the ScurveChartBuilder class."""
    
    @pytest.fixture
    def sample_comparison_data(self):
        """Sample correlation data for testing."""
        return [
            {'Workload': 'BERT', 'Batch': 16, 'Ratio-HLM-to-SiTarget': 0.95},
            {'Workload': 'ResNet', 'Batch': 32, 'Ratio-HLM-to-SiTarget': 1.05},
            {'Workload': 'GPT2', 'Batch': 8, 'Ratio-HLM-to-SiTarget': 0.85},
            {'Workload': 'VGG', 'Batch': 64, 'Ratio-HLM-to-SiTarget': 1.15},
            {'Workload': 'Transformer', 'Batch': 24, 'Ratio-HLM-to-SiTarget': 1.00},
        ]
    
    def test_initialization(self, sample_comparison_data):
        """Test that builder initializes correctly."""
        builder = ScurveChartBuilder(sample_comparison_data)
        
        assert builder.comparison == sample_comparison_data
        assert builder.ratio_data == []
        assert builder.statistics == {}
        assert builder.ws is None
        assert builder.chart is None
    
    def test_prepare_data(self, sample_comparison_data):
        """Test data preparation and sorting."""
        builder = ScurveChartBuilder(sample_comparison_data)
        result = builder.prepare_data()
        
        # Check method chaining
        assert result is builder
        
        # Check data extracted
        assert len(builder.ratio_data) == 5
        
        # Check sorting (should be: 0.85, 0.95, 1.00, 1.05, 1.15)
        ratios = [d['Ratio'] for d in builder.ratio_data]
        assert ratios == [0.85, 0.95, 1.00, 1.05, 1.15]
        
        # Check structure
        first_item = builder.ratio_data[0]
        assert first_item['Workload'] == 'GPT2'
        assert first_item['Batch'] == 8
        assert first_item['Ratio'] == 0.85
    
    def test_prepare_data_filters_invalid_values(self):
        """Test that prepare_data filters out invalid ratio values."""
        data = [
            {'Workload': 'A', 'Batch': 1, 'Ratio-HLM-to-SiTarget': 0.95},
            {'Workload': 'B', 'Batch': 2, 'Ratio-HLM-to-SiTarget': None},
            {'Workload': 'C', 'Batch': 3, 'Ratio-HLM-to-SiTarget': 'invalid'},
            {'Workload': 'D', 'Batch': 4},  # Missing ratio key
            {'Workload': 'E', 'Batch': 5, 'Ratio-HLM-to-SiTarget': 1.05},
        ]
        
        builder = ScurveChartBuilder(data)
        builder.prepare_data()
        
        # Should only have 2 valid entries
        assert len(builder.ratio_data) == 2
        ratios = [d['Ratio'] for d in builder.ratio_data]
        assert ratios == [0.95, 1.05]
    
    def test_calculate_statistics(self, sample_comparison_data):
        """Test statistics calculation."""
        builder = ScurveChartBuilder(sample_comparison_data)
        builder.prepare_data()
        result = builder.calculate_statistics()
        
        # Check method chaining
        assert result is builder
        
        # Check statistics
        assert builder.statistics['min'] == 0.85
        assert builder.statistics['max'] == 1.15
        assert builder.statistics['median'] == 1.00
        
        # Check geometric mean (exp(mean(log(values))))
        ratios = [0.85, 0.95, 1.00, 1.05, 1.15]
        expected_geomean = math.exp(sum(math.log(x) for x in ratios) / len(ratios))
        assert abs(builder.statistics['geomean'] - expected_geomean) < 0.0001
    
    def test_calculate_statistics_raises_on_empty_data(self):
        """Test that calculate_statistics raises error when no data."""
        builder = ScurveChartBuilder([])
        builder.prepare_data()  # Empty ratio_data
        
        with pytest.raises(ValueError, match='Cannot calculate statistics'):
            builder.calculate_statistics()
    
    def test_create_worksheet(self, sample_comparison_data):
        """Test worksheet creation."""
        builder = ScurveChartBuilder(sample_comparison_data)
        builder.prepare_data().calculate_statistics()
        
        wb = Workbook()
        result = builder.create_worksheet(wb)
        
        # Check method chaining
        assert result is builder
        
        # Check worksheet created
        assert builder.ws is not None
        assert builder.ws.title == 'S-Curve Analysis'
        
        # Check title
        assert builder.ws['A1'].value == 'S-Curve Analysis: HLM vs Silicon Target Correlation'
        
        # Check statistics summary
        assert builder.ws['A3'].value == 'Statistics Summary'
        assert builder.ws['B4'].value == 5  # Total workloads
        assert builder.ws['B5'].value == 0.85  # Min ratio
        assert builder.ws['B6'].value == 1.15  # Max ratio
        assert builder.ws['B7'].value == 1.00  # Median ratio
    
    def test_add_data_table(self, sample_comparison_data):
        """Test data table addition."""
        builder = ScurveChartBuilder(sample_comparison_data)
        builder.prepare_data().calculate_statistics()
        
        wb = Workbook()
        builder.create_worksheet(wb)
        result = builder.add_data_table()
        
        # Check method chaining
        assert result is builder
        
        # Check headers
        assert builder.ws['A10'].value == 'Index'
        assert builder.ws['B10'].value == 'Workload'
        assert builder.ws['C10'].value == 'Batch'
        assert builder.ws['D10'].value == 'Ratio-HLM-to-SiTarget'
        
        # Check first data row (sorted, so GPT2 with 0.85 ratio)
        assert builder.ws['A11'].value == 1  # Index
        assert builder.ws['B11'].value == 'GPT2'
        assert builder.ws['C11'].value == 8
        assert builder.ws['D11'].value == 0.85
        
        # Check last data row (VGG with 1.15 ratio)
        assert builder.ws['A15'].value == 5
        assert builder.ws['B15'].value == 'VGG'
        assert builder.ws['D15'].value == 1.15
    
    def test_add_data_table_raises_without_worksheet(self, sample_comparison_data):
        """Test that add_data_table raises error if worksheet not created."""
        builder = ScurveChartBuilder(sample_comparison_data)
        builder.prepare_data()
        
        with pytest.raises(ValueError, match='Worksheet not created'):
            builder.add_data_table()
    
    def test_create_chart(self, sample_comparison_data):
        """Test chart creation."""
        builder = ScurveChartBuilder(sample_comparison_data)
        builder.prepare_data().calculate_statistics()
        
        wb = Workbook()
        builder.create_worksheet(wb).add_data_table()
        result = builder.create_chart()
        
        # Check method chaining
        assert result is builder
        
        # Check chart created
        assert builder.chart is not None
        # chart.title is an object; check it exists
        assert builder.chart.title is not None
        assert builder.chart.style == 12
        assert builder.chart.height == 12
        assert builder.chart.width == 20
        
        # Check axes configured (axes titles are also objects)
        assert builder.chart.x_axis.title is not None
        assert builder.chart.y_axis.title is not None
        
        # Check Y-axis scaling
        assert builder.chart.y_axis.scaling.min >= 0.0
        assert builder.chart.y_axis.scaling.min < 0.85
        assert builder.chart.y_axis.scaling.max > 1.15
    
    def test_create_chart_raises_without_worksheet(self, sample_comparison_data):
        """Test that create_chart raises error if worksheet not created."""
        builder = ScurveChartBuilder(sample_comparison_data)
        builder.prepare_data().calculate_statistics()
        
        with pytest.raises(ValueError, match='Worksheet not created'):
            builder.create_chart()
    
    def test_build_complete_workflow(self, sample_comparison_data):
        """Test the complete build workflow."""
        builder = ScurveChartBuilder(sample_comparison_data)
        wb = Workbook()
        
        # Execute build
        builder.build(wb)
        
        # Check all components created
        assert builder.ratio_data is not None
        assert len(builder.ratio_data) == 5
        assert builder.statistics is not None
        assert builder.ws is not None
        assert builder.chart is not None
        
        # Check worksheet exists in workbook
        assert 'S-Curve Analysis' in wb.sheetnames
        
        # Check column widths adjusted
        assert builder.ws.column_dimensions['A'].width == 8
        assert builder.ws.column_dimensions['B'].width == 30
        assert builder.ws.column_dimensions['C'].width == 10
        assert builder.ws.column_dimensions['D'].width == 20
    
    def test_build_with_empty_data(self):
        """Test build with no valid ratio data."""
        builder = ScurveChartBuilder([])
        wb = Workbook()
        
        # Should handle gracefully (prepare_data returns early if no data)
        # The build process should stop gracefully when no ratio data
        with pytest.raises(ValueError, match='Cannot calculate statistics'):
            builder.build(wb)
    
    def test_method_chaining(self, sample_comparison_data):
        """Test fluent interface (method chaining)."""
        builder = ScurveChartBuilder(sample_comparison_data)
        wb = Workbook()
        
        # Test chaining all methods except build()
        result = (builder
                  .prepare_data()
                  .calculate_statistics()
                  .create_worksheet(wb)
                  .add_data_table()
                  .create_chart())
        
        assert result is builder


class TestAddScurveChart:
    """Tests for the add_scurve_chart function."""
    
    def test_add_scurve_chart_basic(self):
        """Test basic add_scurve_chart functionality."""
        wb = Workbook()
        comparison = [
            {'Workload': 'BERT', 'Batch': 16, 'Ratio-HLM-to-SiTarget': 0.95},
            {'Workload': 'ResNet', 'Batch': 32, 'Ratio-HLM-to-SiTarget': 1.05},
        ]
        
        # Should create sheet without errors
        add_scurve_chart(wb, comparison)
        
        # Check sheet was created
        assert 'S-Curve Analysis' in wb.sheetnames
        ws = wb['S-Curve Analysis']
        
        # Check title exists
        assert ws['A1'].value == 'S-Curve Analysis: HLM vs Silicon Target Correlation'
        
        # Check data exists
        assert ws['A10'].value == 'Index'
        assert ws['A11'].value == 1
        assert ws['A12'].value == 2
    
    def test_add_scurve_chart_uses_builder_internally(self):
        """Test that add_scurve_chart uses ScurveChartBuilder."""
        wb = Workbook()
        comparison = [
            {'Workload': 'A', 'Batch': 1, 'Ratio-HLM-to-SiTarget': 0.9},
            {'Workload': 'B', 'Batch': 2, 'Ratio-HLM-to-SiTarget': 1.1},
        ]
        
        add_scurve_chart(wb, comparison)
        
        # Verify the sheet was created with expected format
        ws = wb['S-Curve Analysis']
        assert ws is not None
        assert ws['A3'].value == 'Statistics Summary'
        assert ws['B10'].value == 'Workload'

